"""
API endpoints с валидацией Pydantic для управления коллекцией штаммов
"""

from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.db import transaction, IntegrityError
from django.db import models
from pydantic import BaseModel, Field, ValidationError, field_validator
from typing import Optional
import logging
from django.db.models import Q
from django.db.models.functions import Lower
from django.db import connection
import re
from django.core.exceptions import ValidationError as DjangoValidationError

from .models import (
    IndexLetter, Location, Source, Comment, AppendixNote,
    Storage, Strain, Sample, StorageBox, SamplePhoto
)
from .schemas import (
    IndexLetterSchema, LocationSchema, SourceSchema,
    CommentSchema, AppendixNoteSchema, StorageSchema,
    StrainSchema, SampleSchema
)
from .utils import log_change, generate_batch_id, model_to_dict, reset_sequence

logger = logging.getLogger(__name__)


@api_view(['GET'])
def api_status(request):
    """Статус API с информацией о валидации"""
    return Response({
        'status': 'OK',
        'validation': 'Pydantic 2.x',
        'endpoints': [
            '/api/strains/',
            '/api/strains/create/',
            '/api/strains/<id>/',
            '/api/strains/<id>/update/',
            '/api/strains/<id>/delete/',
            '/api/strains/validate/',
            '/api/strains/bulk-delete/',
            '/api/strains/bulk-update/',
            '/api/strains/export/',
            '/api/samples/',
            '/api/samples/create/',
            '/api/samples/<id>/',
            '/api/samples/<id>/update/',
            '/api/samples/<id>/delete/',
            '/api/samples/validate/',
            '/api/samples/bulk-delete/',
            '/api/samples/bulk-update/',
            '/api/samples/export/',
            '/api/storage/',
            '/api/stats/',
            '/api/reference-data/',
            # Новые endpoints для боксов
            '/api/reference-data/boxes/',
            '/api/reference-data/boxes/create/',
            '/api/reference-data/boxes/<box_id>/',
            '/api/reference-data/boxes/<box_id>/update/',
            '/api/reference-data/boxes/<box_id>/delete/',
            '/api/reference-data/boxes/<box_id>/cells/',
            '/api/reference-data/boxes/<box_id>/cells/<cell_id>/assign/',
            '/api/reference-data/boxes/<box_id>/cells/<cell_id>/clear/',
            '/api/reference-data/boxes/<box_id>/cells/bulk-assign/',
        ]
    })


@api_view(['GET'])
def list_strains(request):
    """Список всех штаммов с расширенным поиском и фильтрацией - ОПТИМИЗИРОВАННАЯ ВЕРСИЯ"""
    page = max(1, int(request.GET.get('page', 1)))
    limit = max(1, min(int(request.GET.get('limit', 50)), 1000))
    offset = (page - 1) * limit
    
    # Полнотекстовый поиск
    search_query = request.GET.get('search', '').strip()
    
    # Базовые параметры для SQL
    sql_params = []
    where_conditions = []
    
    # Условия поиска
    if search_query:
        search_condition = """(
            st.short_code ILIKE %s OR 
            st.identifier ILIKE %s OR 
            st.rrna_taxonomy ILIKE %s OR 
            st.name_alt ILIKE %s OR 
            st.rcam_collection_id ILIKE %s
        )"""
        where_conditions.append(search_condition)
        search_param = f"%{search_query}%"
        sql_params.extend([search_param] * 5)
    
    # Обработка расширенных фильтров
    def process_filter_param(param_name, param_value):
        """Обработка расширенных операторов фильтрации"""
        if not param_value:
            return
        
        # Карта полей к колонкам в БД
        field_mapping = {
            'short_code': 'st.short_code',
            'identifier': 'st.identifier',
            'rrna_taxonomy': 'st.rrna_taxonomy',
            'rcam_collection_id': 'st.rcam_collection_id',
            'created_at': 'st.created_at'
        }
        
        # Разбираем параметр на поле и оператор
        if '__' in param_name:
            field, operator = param_name.split('__', 1)
        else:
            field = param_name
            # Для числовых/датовых полей безопаснее по умолчанию использовать '='
            numeric_or_date_fields = {'strain_id', 'box_id', 'created_at'}
            operator = 'equals' if field in numeric_or_date_fields else 'ilike'
        
        # Получаем имя колонки
        db_field = field_mapping.get(field)
        if not db_field:
            return
        
        # Применяем оператор
        if operator == 'startswith':
            where_conditions.append(f"{db_field} ILIKE %s")
            sql_params.append(f"{param_value}%")
        elif operator == 'endswith':
            where_conditions.append(f"{db_field} ILIKE %s")
            sql_params.append(f"%{param_value}")
        elif operator == 'contains' or operator == 'ilike':
            where_conditions.append(f"{db_field} ILIKE %s")
            sql_params.append(f"%{param_value}%")
        elif operator == 'equals' or operator == 'exact':
            if field == 'created_at':
                # Для дат используем точное сравнение
                try:
                    from datetime import datetime
                    date_value = datetime.fromisoformat(param_value.replace('Z', '+00:00'))
                    where_conditions.append(f"{db_field} = %s")
                    sql_params.append(date_value)
                except ValueError:
                    pass
            else:
                # Приводим к числу, если поле числовое
                if field == 'strain_id':
                    try:
                        param_value = int(param_value)
                    except ValueError:
                        return
                where_conditions.append(f"{db_field} = %s")
                sql_params.append(param_value)
        elif operator == 'gt':
            if field == 'created_at':
                try:
                    from datetime import datetime
                    date_value = datetime.fromisoformat(param_value.replace('Z', '+00:00'))
                    where_conditions.append(f"{db_field} > %s")
                    sql_params.append(date_value)
                except ValueError:
                    pass
            else:
                where_conditions.append(f"{db_field} > %s")
                sql_params.append(param_value)
        elif operator == 'lt':
            if field == 'created_at':
                try:
                    from datetime import datetime
                    date_value = datetime.fromisoformat(param_value.replace('Z', '+00:00'))
                    where_conditions.append(f"{db_field} < %s")
                    sql_params.append(date_value)
                except ValueError:
                    pass
            else:
                where_conditions.append(f"{db_field} < %s")
                sql_params.append(param_value)
        elif operator == 'gte':
            if field == 'created_at':
                try:
                    from datetime import datetime
                    date_value = datetime.fromisoformat(param_value.replace('Z', '+00:00'))
                    where_conditions.append(f"{db_field} >= %s")
                    sql_params.append(date_value)
                except ValueError:
                    pass
            else:
                where_conditions.append(f"{db_field} >= %s")
                sql_params.append(param_value)
        elif operator == 'lte':
            if field == 'created_at':
                try:
                    from datetime import datetime
                    date_value = datetime.fromisoformat(param_value.replace('Z', '+00:00'))
                    where_conditions.append(f"{db_field} <= %s")
                    sql_params.append(date_value)
                except ValueError:
                    pass
            else:
                where_conditions.append(f"{db_field} <= %s")
                sql_params.append(param_value)

    # Обрабатываем все GET параметры
    for param_name, param_value in request.GET.items():
        if param_name in ['page', 'limit', 'search']:
            continue  # Пропускаем системные параметры
        
        # Обратная совместимость для старых параметров
        if param_name == 'rcam_id':
            process_filter_param('rcam_collection_id', param_value)
        elif param_name == 'taxonomy':
            process_filter_param('rrna_taxonomy', param_value)
        elif param_name == 'created_after':
            process_filter_param('created_at__gte', param_value)
        elif param_name == 'created_before':
            process_filter_param('created_at__lte', param_value)
        else:
            # Новые расширенные фильтры
            process_filter_param(param_name, param_value)
    
    # Формируем WHERE условие
    where_clause = ""
    if where_conditions:
        where_clause = "WHERE " + " AND ".join(where_conditions)
    
    # SQL для подсчета общего количества
    count_sql = f"""
        SELECT COUNT(*) as total
        FROM collection_manager_strain st
        {where_clause}
    """
    
    # SQL для получения данных с пагинацией
    data_sql = f"""
        SELECT 
            st.id,
            st.short_code,
            st.identifier,
            st.rrna_taxonomy,
            st.name_alt,
            st.rcam_collection_id,
            st.created_at,
            st.updated_at
        FROM collection_manager_strain st
        {where_clause}
        ORDER BY st.id
        LIMIT %s OFFSET %s
    """
    
    with connection.cursor() as cursor:
        # Получаем общее количество
        cursor.execute(count_sql, sql_params)
        total_count = cursor.fetchone()[0]
        
        # Получаем данные
        cursor.execute(data_sql, sql_params + [limit, offset])
        columns = [col[0] for col in cursor.description]
        strains_data = [dict(zip(columns, row)) for row in cursor.fetchall()]
    
    # Метаданные пагинации
    total_pages = (total_count + limit - 1) // limit
    has_next = page < total_pages
    has_previous = page > 1
    
    # Форматируем данные
    data = []
    for strain in strains_data:
        data.append({
            'id': strain['id'],
            'short_code': strain['short_code'],
            'identifier': strain['identifier'],
            'rrna_taxonomy': strain['rrna_taxonomy'],
            'name_alt': strain['name_alt'],
            'rcam_collection_id': strain['rcam_collection_id'],
            'created_at': strain['created_at'].isoformat() if strain['created_at'] else None,
            'updated_at': strain['updated_at'].isoformat() if strain['updated_at'] else None,
        })
    
    return Response({
        'strains': data,
        'pagination': {
            'total': total_count,
            'shown': len(data),
            'page': page,
            'limit': limit,
            'total_pages': total_pages,
            'has_next': has_next,
            'has_previous': has_previous,
            'offset': offset
        },
        'search_query': search_query,
        'filters_applied': {
            'search': bool(search_query),
            'advanced_filters': [param for param in request.GET.keys() 
                               if param not in ['page', 'limit', 'search'] and request.GET[param]],
            'total_filters': len([param for param in request.GET.keys() 
                                if param not in ['page', 'limit', 'search'] and request.GET[param]]),
        }
    })


@api_view(['GET'])
def get_strain(request, strain_id):
    """Получение детальной информации о штамме"""
    try:
        strain = Strain.objects.select_related().prefetch_related('samples').get(id=strain_id)
        
        # Подсчитываем статистику по образцам
        samples_count = strain.samples.count()
        samples_with_photo = strain.samples.filter(has_photo=True).count()
        samples_identified = strain.samples.filter(is_identified=True).count()
        samples_with_genome = strain.samples.filter(has_genome=True).count()
        samples_with_biochemistry = strain.samples.filter(has_biochemistry=True).count()
        
        return Response({
            'id': strain.id,
            'short_code': strain.short_code,
            'identifier': strain.identifier,
            'rrna_taxonomy': strain.rrna_taxonomy,
            'name_alt': strain.name_alt,
            'rcam_collection_id': strain.rcam_collection_id,
            'created_at': strain.created_at.isoformat() if strain.created_at else None,
            'updated_at': strain.updated_at.isoformat() if strain.updated_at else None,
            # Статистика по образцам
            'samples_stats': {
                'total_count': samples_count,
                'with_photo_count': samples_with_photo,
                'identified_count': samples_identified,
                'with_genome_count': samples_with_genome,
                'with_biochemistry_count': samples_with_biochemistry,
                'photo_percentage': round((samples_with_photo / samples_count * 100) if samples_count > 0 else 0, 1),
                'identified_percentage': round((samples_identified / samples_count * 100) if samples_count > 0 else 0, 1),
            }
        })
    
    except Strain.DoesNotExist:
        return Response({
            'error': 'Штамм не найден'
        }, status=status.HTTP_404_NOT_FOUND)
    
    except Exception as e:
        logger.error(f"Unexpected error in get_strain: {e}")
        return Response({
            'error': 'Внутренняя ошибка сервера'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['PUT', 'PATCH'])
def update_strain(request, strain_id):
    """Обновление существующего штамма"""
    try:
        # Создаем схему для обновления без ID
        class UpdateStrainSchema(BaseModel):
            short_code: str = Field(min_length=1, max_length=50, description="Короткий код штамма")
            rrna_taxonomy: Optional[str] = Field(None, max_length=500, description="rRNA таксономия")
            identifier: str = Field(min_length=1, max_length=100, description="Идентификатор штамма")
            name_alt: Optional[str] = Field(None, max_length=200, description="Альтернативное название")
            rcam_collection_id: Optional[str] = Field(None, max_length=50, description="ID коллекции RCAM")
            
            @field_validator('short_code', 'identifier')
            @classmethod
            def validate_required_fields(cls, v: str) -> str:
                return v.strip()
            
            @field_validator('rrna_taxonomy', 'name_alt', 'rcam_collection_id')
            @classmethod
            def validate_optional_fields(cls, v: Optional[str]) -> Optional[str]:
                if v is not None:
                    v = v.strip()
                    return v if v else None
                return v
        
        # Проверяем существование штамма
        try:
            strain = Strain.objects.get(id=strain_id)
        except Strain.DoesNotExist:
            return Response({
                'error': f'Штамм с ID {strain_id} не найден'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Валидация входных данных
        validated_data = UpdateStrainSchema.model_validate(request.data)
        
        # Проверка уникальности короткого кода (исключая текущий штамм)
        if Strain.objects.filter(short_code=validated_data.short_code).exclude(id=strain_id).exists():
            return Response({
                'error': f'Штамм с кодом "{validated_data.short_code}" уже существует'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Обновление штамма
        with transaction.atomic():
            strain.short_code = validated_data.short_code
            strain.rrna_taxonomy = validated_data.rrna_taxonomy
            strain.identifier = validated_data.identifier
            strain.name_alt = validated_data.name_alt
            strain.rcam_collection_id = validated_data.rcam_collection_id
            strain.save()
            
            logger.info(f"Updated strain: {strain.short_code} (ID: {strain.id})")
        
        return Response({
            'id': strain.id,
            'short_code': strain.short_code,
            'identifier': strain.identifier,
            'rrna_taxonomy': strain.rrna_taxonomy,
            'name_alt': strain.name_alt,
            'rcam_collection_id': strain.rcam_collection_id,
            'message': 'Штамм успешно обновлен'
        })
    
    except ValidationError as e:
        return Response({
            'error': 'Ошибки валидации данных',
            'details': e.errors()
        }, status=status.HTTP_400_BAD_REQUEST)
    
    except Exception as e:
        logger.error(f"Unexpected error in update_strain: {e}")
        return Response({
            'error': 'Внутренняя ошибка сервера'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['DELETE'])
def delete_strain(request, strain_id):
    """Удаление штамма"""
    try:
        # Получаем штамм
        try:
            strain = Strain.objects.get(id=strain_id)
        except Strain.DoesNotExist:
            return Response({
                'error': f'Штамм с ID {strain_id} не найден'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Проверяем, есть ли связанные образцы
        related_samples = Sample.objects.filter(strain_id=strain_id).count()
        if related_samples > 0:
            return Response({
                'error': f'Нельзя удалить штамм, так как с ним связано {related_samples} образцов',
                'related_samples_count': related_samples
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Сохраняем информацию о штамме для ответа
        strain_info = {
            'id': strain.id,
            'short_code': strain.short_code,
            'identifier': strain.identifier
        }
        
        # Удаляем штамм
        with transaction.atomic():
            strain.delete()
            logger.info(f"Deleted strain: {strain_info['short_code']} (ID: {strain_info['id']})")
        
        return Response({
            'message': f"Штамм '{strain_info['short_code']}' успешно удален",
            'deleted_strain': strain_info
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Unexpected error in delete_strain: {e}")
        return Response({
            'error': 'Внутренняя ошибка сервера'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def create_strain(request):
    """Создание нового штамма"""
    try:
        # Создаем схему для создания без ID
        class CreateStrainSchema(BaseModel):
            short_code: str = Field(min_length=1, max_length=50, description="Короткий код штамма")
            rrna_taxonomy: Optional[str] = Field(None, max_length=500, description="rRNA таксономия")
            identifier: str = Field(min_length=1, max_length=100, description="Идентификатор штамма")
            name_alt: Optional[str] = Field(None, max_length=200, description="Альтернативное название")
            rcam_collection_id: Optional[str] = Field(None, max_length=50, description="ID коллекции RCAM")
            
            @field_validator('short_code', 'identifier')
            @classmethod
            def validate_required_fields(cls, v: str) -> str:
                return v.strip()
            
            @field_validator('rrna_taxonomy', 'name_alt', 'rcam_collection_id')
            @classmethod
            def validate_optional_fields(cls, v: Optional[str]) -> Optional[str]:
                if v is not None:
                    v = v.strip()
                    return v if v else None
                return v
        
        # Валидация входных данных
        validated_data = CreateStrainSchema.model_validate(request.data)
        
        # Проверка уникальности короткого кода
        if Strain.objects.filter(short_code=validated_data.short_code).exists():
            return Response({
                'error': f'Штамм с кодом "{validated_data.short_code}" уже существует'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Создание штамма с авто-восстановлением последовательности PK при необходимости
        try:
            with transaction.atomic():
                strain = Strain.objects.create(
                    short_code=validated_data.short_code,
                    rrna_taxonomy=validated_data.rrna_taxonomy,
                    identifier=validated_data.identifier,
                    name_alt=validated_data.name_alt,
                    rcam_collection_id=validated_data.rcam_collection_id,
                )
                logger.info(
                    f"Created new strain: {strain.short_code} (ID: {strain.id})"
                )
        except IntegrityError as ie:
            # Попытка исправить несинхронизированную последовательность PK
            if "duplicate key value violates unique constraint" in str(ie) and "_pkey" in str(ie):
                logger.warning(
                    "IntegrityError on Strain create — attempting to reset sequence and retry"
                )
                reset_sequence(Strain)
                with transaction.atomic():
                    strain = Strain.objects.create(
                        short_code=validated_data.short_code,
                        rrna_taxonomy=validated_data.rrna_taxonomy,
                        identifier=validated_data.identifier,
                        name_alt=validated_data.name_alt,
                        rcam_collection_id=validated_data.rcam_collection_id,
                    )
                    logger.info(
                        f"Created new strain after sequence reset: {strain.short_code} (ID: {strain.id})"
                    )
            else:
                raise
        
        return Response({
            'id': strain.id,
            'short_code': strain.short_code,
            'identifier': strain.identifier,
            'rrna_taxonomy': strain.rrna_taxonomy,
            'name_alt': strain.name_alt,
            'rcam_collection_id': strain.rcam_collection_id,
            'message': 'Штамм успешно создан'
        }, status=status.HTTP_201_CREATED)
    
    except ValidationError as e:
        return Response({
            'error': 'Ошибки валидации данных',
            'details': e.errors()
        }, status=status.HTTP_400_BAD_REQUEST)
    
    except Exception as e:
        logger.error(f"Unexpected error in create_strain: {e}")
        return Response({
            'error': 'Внутренняя ошибка сервера'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def validate_strain(request):
    """Валидация данных штамма с помощью Pydantic"""
    try:
        # Используем схему без ID для валидации данных создания
        class CreateStrainSchema(BaseModel):
            short_code: str = Field(min_length=1, max_length=50, description="Короткий код штамма")
            rrna_taxonomy: Optional[str] = Field(None, max_length=500, description="rRNA таксономия")
            identifier: str = Field(min_length=1, max_length=100, description="Идентификатор штамма")
            name_alt: Optional[str] = Field(None, max_length=200, description="Альтернативное название")
            rcam_collection_id: Optional[str] = Field(None, max_length=50, description="ID коллекции RCAM")
            
            @field_validator('short_code', 'identifier')
            @classmethod
            def validate_required_fields(cls, v: str) -> str:
                return v.strip()
            
            @field_validator('rrna_taxonomy', 'name_alt', 'rcam_collection_id')
            @classmethod
            def validate_optional_fields(cls, v: Optional[str]) -> Optional[str]:
                if v is not None:
                    v = v.strip()
                    return v if v else None
                return v
        
        validated_data = CreateStrainSchema.model_validate(request.data)
        
        return Response({
            'valid': True,
            'data': validated_data.model_dump(),
            'message': 'Данные штамма корректны'
        })
    
    except ValidationError as e:
        return Response({
            'valid': False,
            'errors': e.errors(),
            'message': 'Ошибки валидации данных'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    except Exception as e:
        logger.error(f"Unexpected error in validate_strain: {e}")
        return Response({
            'valid': False,
            'error': str(e),
            'message': 'Внутренняя ошибка сервера'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def list_samples(request):
    """Список образцов с связанными данными и расширенным поиском - ОПТИМИЗИРОВАННАЯ ВЕРСИЯ"""
    page = int(request.GET.get('page', 1))
    limit = min(int(request.GET.get('limit', 100)), 1000)  # Максимум 1000 записей за раз
    offset = (page - 1) * limit
    
    # Полнотекстовый поиск
    search_query = request.GET.get('search', '').strip()
    
    # Базовые параметры для SQL
    sql_params = []
    where_conditions = []
    
    # Условия поиска
    if search_query:
        search_condition = """(
            sam.original_sample_number ILIKE %s OR
            st.short_code ILIKE %s OR
            st.identifier ILIKE %s OR
            st.rrna_taxonomy ILIKE %s OR
            st.name_alt ILIKE %s OR
            st.rcam_collection_id ILIKE %s OR
            src.organism_name ILIKE %s OR
            src.source_type ILIKE %s OR
            src.category ILIKE %s OR
            loc.name ILIKE %s OR
            storage.box_id ILIKE %s OR
            storage.cell_id ILIKE %s OR
            idx.letter_value ILIKE %s
        )"""
        where_conditions.append(search_condition)
        search_param = f"%{search_query}%"
        sql_params.extend([search_param] * 13)
    
    # ------------------ Расширенные фильтры (динамические) ------------------ #
    def process_sample_filter_param(param_name, param_value):
        """Обработка расширенных операторов фильтрации для образцов"""
        if not param_value:
            return

        field_mapping = {
            'original_sample_number': 'sam.original_sample_number',
            'strain_id': 'sam.strain_id',
            'box_id': 'storage.box_id',
            'source_type': 'src.source_type',
            'organism_name': 'src.organism_name',
            'created_at': 'sam.created_at',
        }

        if '__' in param_name:
            field, operator = param_name.split('__', 1)
        else:
            field = param_name
            # Для числовых/датовых полей и полей выбора безопаснее по умолчанию использовать '='
            exact_match_fields = {'strain_id', 'box_id', 'created_at', 'source_type', 'organism_name'}
            operator = 'equals' if field in exact_match_fields else 'ilike'

        db_field = field_mapping.get(field)
        if not db_field:
            return

        # Текстовые и числовые операторы
        if operator == 'startswith':
            where_conditions.append(f"{db_field} ILIKE %s")
            sql_params.append(f"{param_value}%")
        elif operator == 'endswith':
            where_conditions.append(f"{db_field} ILIKE %s")
            sql_params.append(f"%{param_value}")
        elif operator in ['contains', 'ilike']:
            where_conditions.append(f"{db_field} ILIKE %s")
            sql_params.append(f"%{param_value}%")
        elif operator in ['equals', 'exact']:
            if field == 'created_at':
                try:
                    from datetime import datetime
                    date_val = datetime.fromisoformat(param_value.replace('Z', '+00:00'))
                    where_conditions.append(f"{db_field} = %s")
                    sql_params.append(date_val)
                except ValueError:
                    pass
            else:
                # Приводим к числу, если поле числовое
                if field == 'strain_id':
                    try:
                        param_value = int(param_value)
                    except ValueError:
                        return
                where_conditions.append(f"{db_field} = %s")
                sql_params.append(param_value)
        elif operator == 'gt':
            where_conditions.append(f"{db_field} > %s")
            sql_params.append(param_value)
        elif operator == 'lt':
            where_conditions.append(f"{db_field} < %s")
            sql_params.append(param_value)
        elif operator == 'gte':
            where_conditions.append(f"{db_field} >= %s")
            sql_params.append(param_value)
        elif operator == 'lte':
            where_conditions.append(f"{db_field} <= %s")
            sql_params.append(param_value)

    # Перебираем все GET параметры и применяем процессинг
    for p_name, p_val in request.GET.items():
        if p_name in ['page', 'limit', 'search']:
            continue
        process_sample_filter_param(p_name, p_val)
    
    # Дополнительные фильтры
    if 'has_photo' in request.GET:
        where_conditions.append("sam.has_photo = %s")
        sql_params.append(request.GET['has_photo'].lower() == 'true')
        
    if 'is_identified' in request.GET:
        where_conditions.append("sam.is_identified = %s")
        sql_params.append(request.GET['is_identified'].lower() == 'true')
        
    if 'has_antibiotic_activity' in request.GET:
        where_conditions.append("sam.has_antibiotic_activity = %s")
        sql_params.append(request.GET['has_antibiotic_activity'].lower() == 'true')
        
    if 'has_genome' in request.GET:
        where_conditions.append("sam.has_genome = %s")
        sql_params.append(request.GET['has_genome'].lower() == 'true')
        
    if 'has_biochemistry' in request.GET:
        where_conditions.append("sam.has_biochemistry = %s")
        sql_params.append(request.GET['has_biochemistry'].lower() == 'true')
        
    if 'seq_status' in request.GET:
        where_conditions.append("sam.seq_status = %s")
        sql_params.append(request.GET['seq_status'].lower() == 'true')
        
    if 'source_id' in request.GET:
        where_conditions.append("sam.source_id = %s")
        sql_params.append(int(request.GET['source_id']))
        
    if 'location_id' in request.GET:
        where_conditions.append("sam.location_id = %s")
        sql_params.append(int(request.GET['location_id']))
        

        
    # Фильтры по датам
    if 'created_after' in request.GET:
        try:
            from datetime import datetime
            date_after = datetime.fromisoformat(request.GET['created_after'].replace('Z', '+00:00'))
            where_conditions.append("sam.created_at >= %s")
            sql_params.append(date_after)
        except ValueError:
            pass
            
    if 'created_before' in request.GET:
        try:
            from datetime import datetime
            date_before = datetime.fromisoformat(request.GET['created_before'].replace('Z', '+00:00'))
            where_conditions.append("sam.created_at <= %s")
            sql_params.append(date_before)
        except ValueError:
            pass
    
    # Формируем WHERE условие
    where_clause = ""
    if where_conditions:
        where_clause = "WHERE " + " AND ".join(where_conditions)
    
    # SQL для подсчета общего количества
    count_sql = f"""
        SELECT COUNT(*) as total
        FROM collection_manager_sample sam
        LEFT JOIN collection_manager_strain st ON sam.strain_id = st.id
        LEFT JOIN collection_manager_storage storage ON sam.storage_id = storage.id
        LEFT JOIN collection_manager_source src ON sam.source_id = src.id
        LEFT JOIN collection_manager_location loc ON sam.location_id = loc.id
        LEFT JOIN collection_manager_indexletter idx ON sam.index_letter_id = idx.id
        {where_clause}
    """
    
    # SQL для получения данных с пагинацией и всеми связанными данными
    data_sql = f"""
        SELECT 
            sam.id,
            sam.original_sample_number,
            sam.has_photo,
            sam.is_identified,
            sam.has_antibiotic_activity,
            sam.has_genome,
            sam.has_biochemistry,
            sam.seq_status,
            sam.created_at,
            sam.updated_at,
            -- Strain data
            st.id as strain_id,
            st.short_code as strain_short_code,
            st.identifier as strain_identifier,
            st.rrna_taxonomy as strain_rrna_taxonomy,
            -- Storage data
            storage.id as storage_id,
            storage.box_id as storage_box_id,
            storage.cell_id as storage_cell_id,
            -- Source data
            src.id as source_id,
            src.organism_name as source_organism_name,
            src.source_type as source_source_type,
            src.category as source_category,
            -- Location data
            loc.id as location_id,
            loc.name as location_name,
            -- Index letter data
            idx.id as index_letter_id,
            idx.letter_value as index_letter_value
        FROM collection_manager_sample sam
        LEFT JOIN collection_manager_strain st ON sam.strain_id = st.id
        LEFT JOIN collection_manager_storage storage ON sam.storage_id = storage.id
        LEFT JOIN collection_manager_source src ON sam.source_id = src.id
        LEFT JOIN collection_manager_location loc ON sam.location_id = loc.id
        LEFT JOIN collection_manager_indexletter idx ON sam.index_letter_id = idx.id
        {where_clause}
        ORDER BY sam.id
        LIMIT %s OFFSET %s
    """
    
    with connection.cursor() as cursor:
        # Получаем общее количество
        cursor.execute(count_sql, sql_params)
        total_count = cursor.fetchone()[0]
        
        # Получаем данные
        cursor.execute(data_sql, sql_params + [limit, offset])
        columns = [col[0] for col in cursor.description]
        samples_data = [dict(zip(columns, row)) for row in cursor.fetchall()]
    
    # Вычисляем метаданные пагинации
    total_pages = (total_count + limit - 1) // limit  # Округление вверх
    has_next = page < total_pages
    has_previous = page > 1
    
    # Форматируем данные
    data = []
    for sample in samples_data:
        data.append({
            'id': sample['id'],
            'strain': {
                'id': sample['strain_id'],
                'short_code': sample['strain_short_code'],
                'identifier': sample['strain_identifier'],
                'rrna_taxonomy': sample['strain_rrna_taxonomy'],
            } if sample['strain_id'] else None,
            'storage': {
                'id': sample['storage_id'],
                'box_id': sample['storage_box_id'],
                'cell_id': sample['storage_cell_id'],
            } if sample['storage_id'] else None,
            'source': {
                'id': sample['source_id'],
                'organism_name': sample['source_organism_name'],
                'source_type': sample['source_source_type'],
                'category': sample['source_category'],
            } if sample['source_id'] else None,
            'location': {
                'id': sample['location_id'],
                'name': sample['location_name'],
            } if sample['location_id'] else None,
            'index_letter': {
                'id': sample['index_letter_id'],
                'letter_value': sample['index_letter_value'],
            } if sample['index_letter_id'] else None,
            'original_sample_number': sample['original_sample_number'],
            'has_photo': sample['has_photo'],
            'is_identified': sample['is_identified'],
            'has_antibiotic_activity': sample['has_antibiotic_activity'],
            'has_genome': sample['has_genome'],
            'has_biochemistry': sample['has_biochemistry'],
            'seq_status': sample['seq_status'],
            'created_at': sample['created_at'].isoformat() if sample['created_at'] else None,
            'updated_at': sample['updated_at'].isoformat() if sample['updated_at'] else None,
        })
    
    return Response({
        'samples': data,
        'pagination': {
            'total': total_count,
            'shown': len(data),
            'page': page,
            'limit': limit,
            'total_pages': total_pages,
            'has_next': has_next,
            'has_previous': has_previous,
            'offset': offset
        },
        'search_query': search_query,
        'filters_applied': {
            'search': bool(search_query),
            'strain_id': 'strain_id' in request.GET,
            'has_photo': 'has_photo' in request.GET,
            'is_identified': 'is_identified' in request.GET,
            'has_antibiotic_activity': 'has_antibiotic_activity' in request.GET,
            'has_genome': 'has_genome' in request.GET,
            'has_biochemistry': 'has_biochemistry' in request.GET,
            'seq_status': 'seq_status' in request.GET,
            'box_id': 'box_id' in request.GET,
            'source_id': 'source_id' in request.GET,
            'location_id': 'location_id' in request.GET,
            'date_range': 'created_after' in request.GET or 'created_before' in request.GET,
            'advanced_filters': [param for param in request.GET.keys() if param not in ['page', 'limit', 'search'] and request.GET[param]],
            'total_filters': len([param for param in request.GET.keys() if param not in ['page', 'limit', 'search'] and request.GET[param]]),
        }
    })


@api_view(['POST'])
def validate_sample(request):
    """Валидация данных образца с помощью Pydantic"""
    try:
        validated_data = SampleSchema.model_validate(request.data)
        
        return Response({
            'valid': True,
            'data': validated_data.model_dump(),
            'message': 'Данные образца корректны'
        })
    
    except ValidationError as e:
        return Response({
            'valid': False,
            'errors': e.errors(),
            'message': 'Ошибки валидации данных'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    except Exception as e:
        logger.error(f"Unexpected error in validate_sample: {e}")
        return Response({
            'valid': False,
            'error': str(e),
            'message': 'Внутренняя ошибка сервера'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def list_storage(request):
    """Информация о хранилищах с заполненностью"""
    storage_data = {}
    
    for storage in Storage.objects.all():
        box_id = storage.box_id
        if box_id not in storage_data:
            storage_data[box_id] = {
                'box_id': box_id,
                'cells': [],
                'occupied': 0,
                'total': 0
            }
        
        # Получаем сначала «нормальный» образец (не помеченный как свободная ячейка)
        sample = (
            storage.sample_set.select_related('strain', 'comment')
            .exclude(comment_id=2)
            .filter(strain_id__isnull=False)
            .first()
        )
        # Если таких нет — берём первый образец с пометкой «свободная ячейка»
        free_sample = None
        if sample is None:
            free_sample = (
                storage.sample_set.select_related('comment')
                .filter(comment_id=2)
                .first()
            )
        
        # Логика занятости:
        # 1) есть «нормальный» образец с привязанным штаммом → занято
        # 2) иначе — свободно (либо нет образцов, либо помечено как свободная ячейка)
        is_occupied = sample is not None and sample.strain is not None
        actual_sample = sample or free_sample  # то, что вернём в ответе (None, если ячейка полностью пустая)
        
        cell_info = {
            'cell_id': storage.cell_id,
            'storage_id': storage.id,
            'occupied': is_occupied,
            'sample_id': actual_sample.id if actual_sample else None,
            'strain_code': actual_sample.strain.short_code if actual_sample and actual_sample.strain else None,
            # Свободной считаем ячейку, если явно помечена comment_id=2 ИЛИ совсем пустая (нет образцов)
            'is_free_cell': (free_sample is not None) or (actual_sample is None)
        }
        
        storage_data[box_id]['cells'].append(cell_info)
        storage_data[box_id]['total'] += 1
        if is_occupied:
            storage_data[box_id]['occupied'] += 1
    
    # Преобразуем в список боксов
    boxes = list(storage_data.values())
    
    # Сортируем ячейки для каждого бокса
    for box in boxes:
        box['cells'].sort(key=lambda x: x['cell_id'])
    
    # Натуральная сортировка: если в ID есть число — сортируем по числу, иначе по строке
    def _box_sort_key(item):
        m = re.search(r"(\d+)$", str(item['box_id']))
        if m:
            return (0, int(m.group(1)))  # сначала числовые, по возрастанию
        return (1, str(item['box_id']))  # затем остальные лексикографически

    boxes.sort(key=_box_sort_key)
    
    # Вычисляем общую статистику
    total_boxes = len(boxes)
    total_cells = sum(box['total'] for box in boxes)
    occupied_cells = sum(box['occupied'] for box in boxes)
    
    return Response({
        'boxes': boxes,
        'total_boxes': total_boxes,
        'total_cells': total_cells,
        'occupied_cells': occupied_cells
    })


@api_view(['GET'])
def api_stats(request):
    """Общая статистика с информацией о валидации"""
    from django.db import connection
    
    # Получаем правильное количество ячеек хранения (как в других API)
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                s.box_id,
                COUNT(*) as total_cells
            FROM collection_manager_storage s
            LEFT JOIN collection_manager_sample sam ON s.id = sam.storage_id
            GROUP BY s.box_id
        """)
        storage_rows = cursor.fetchall()
        total_storage_cells = sum(row[1] for row in storage_rows)
    
    return Response({
        'counts': {
            'strains': Strain.objects.count(),
            'samples': Sample.objects.count(),
            'storage_units': total_storage_cells,  # ИСПРАВЛЕНО: используем правильный подсчет
            'sources': Source.objects.count(),
            'locations': Location.objects.count(),
            'index_letters': IndexLetter.objects.count(),
            'comments': Comment.objects.count(),
            'appendix_notes': AppendixNote.objects.count(),
        },
        'samples_analysis': {
            'with_photo': Sample.objects.filter(has_photo=True).count(),
            'identified': Sample.objects.filter(is_identified=True).count(),
            'with_antibiotic_activity': Sample.objects.filter(has_antibiotic_activity=True).count(),
            'with_genome': Sample.objects.filter(has_genome=True).count(),
            'with_biochemistry': Sample.objects.filter(has_biochemistry=True).count(),
            'sequenced': Sample.objects.filter(seq_status=True).count(),
        },
        'validation': {
            'engine': 'Pydantic 2.x',
            'schemas_available': [
                'StrainSchema', 'SampleSchema', 'StorageSchema',
                'SourceSchema', 'LocationSchema', 'IndexLetterSchema',
                'CommentSchema', 'AppendixNoteSchema'
            ]
        }
    })


@api_view(['POST'])
def create_sample(request):
    """Создание нового образца"""
    try:
        # Динамически создаем схему валидации
        class CreateSampleSchema(BaseModel):
            strain_id: Optional[int] = Field(None, ge=1, description="ID штамма")
            index_letter_id: Optional[int] = Field(None, ge=1, description="ID индексной буквы")
            storage_id: Optional[int] = Field(None, ge=1, description="ID хранилища")
            original_sample_number: Optional[str] = Field(None, max_length=100, description="Оригинальный номер образца")
            source_id: Optional[int] = Field(None, ge=1, description="ID источника")
            location_id: Optional[int] = Field(None, ge=1, description="ID местоположения")
            appendix_note_id: Optional[int] = Field(None, ge=1, description="ID примечания")
            comment_id: Optional[int] = Field(None, ge=1, description="ID комментария")
            has_photo: bool = Field(default=False, description="Есть ли фото")
            is_identified: bool = Field(default=False, description="Идентифицирован ли")
            has_antibiotic_activity: bool = Field(default=False, description="Есть ли антибиотическая активность")
            has_genome: bool = Field(default=False, description="Есть ли геном")
            has_biochemistry: bool = Field(default=False, description="Есть ли биохимия")
            seq_status: bool = Field(default=False, description="Статус секвенирования")
            
            @field_validator('original_sample_number')
            @classmethod
            def validate_sample_number(cls, v: Optional[str]) -> Optional[str]:
                if v is not None:
                    v = v.strip()
                    return v if v else None
                return v

        validated_data = CreateSampleSchema.model_validate(request.data)
        
        # Проверяем существование связанных объектов
        if validated_data.strain_id:
            if not Strain.objects.filter(id=validated_data.strain_id).exists():
                return Response({
                    'error': f'Штамм с ID {validated_data.strain_id} не найден'
                }, status=status.HTTP_404_NOT_FOUND)
        
        if validated_data.storage_id:
            if not Storage.objects.filter(id=validated_data.storage_id).exists():
                return Response({
                    'error': f'Хранилище с ID {validated_data.storage_id} не найдено'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Проверяем, не занята ли ячейка
            existing_sample = Sample.objects.filter(
                storage_id=validated_data.storage_id
            ).exclude(comment_id=2).first()  # Исключаем помеченные как свободные
            
            if existing_sample:
                storage = Storage.objects.get(id=validated_data.storage_id)
                return Response({
                    'error': f'Ячейка {storage.box_id}-{storage.cell_id} уже занята образцом {existing_sample.id}'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        if validated_data.source_id:
            if not Source.objects.filter(id=validated_data.source_id).exists():
                return Response({
                    'error': f'Источник с ID {validated_data.source_id} не найден'
                }, status=status.HTTP_404_NOT_FOUND)
        
        if validated_data.location_id:
            if not Location.objects.filter(id=validated_data.location_id).exists():
                return Response({
                    'error': f'Местоположение с ID {validated_data.location_id} не найдено'
                }, status=status.HTTP_404_NOT_FOUND)
        
        if validated_data.index_letter_id:
            if not IndexLetter.objects.filter(id=validated_data.index_letter_id).exists():
                return Response({
                    'error': f'Индексная буква с ID {validated_data.index_letter_id} не найдена'
                }, status=status.HTTP_404_NOT_FOUND)
        
        if validated_data.appendix_note_id:
            if not AppendixNote.objects.filter(id=validated_data.appendix_note_id).exists():
                return Response({
                    'error': f'Примечание с ID {validated_data.appendix_note_id} не найдено'
                }, status=status.HTTP_404_NOT_FOUND)
        
        if validated_data.comment_id:
            if not Comment.objects.filter(id=validated_data.comment_id).exists():
                return Response({
                    'error': f'Комментарий с ID {validated_data.comment_id} не найден'
                }, status=status.HTTP_404_NOT_FOUND)
        
        # Создание образца с авто-сбросом последовательности при необходимости
        try:
            with transaction.atomic():
                sample = Sample.objects.create(
                    strain_id=validated_data.strain_id,
                    index_letter_id=validated_data.index_letter_id,
                    storage_id=validated_data.storage_id,
                    original_sample_number=validated_data.original_sample_number,
                    source_id=validated_data.source_id,
                    location_id=validated_data.location_id,
                    appendix_note_id=validated_data.appendix_note_id,
                    comment_id=validated_data.comment_id,
                    has_photo=validated_data.has_photo,
                    is_identified=validated_data.is_identified,
                    has_antibiotic_activity=validated_data.has_antibiotic_activity,
                    has_genome=validated_data.has_genome,
                    has_biochemistry=validated_data.has_biochemistry,
                    seq_status=validated_data.seq_status,
                )
                logger.info(f"Created new sample (ID: {sample.id})")
        except IntegrityError as ie:
            if "duplicate key value violates unique constraint" in str(ie) and "_pkey" in str(ie):
                logger.warning("IntegrityError on Sample create — resetting sequence and retry")
                reset_sequence(Sample)
                with transaction.atomic():
                    sample = Sample.objects.create(
                        strain_id=validated_data.strain_id,
                        index_letter_id=validated_data.index_letter_id,
                        storage_id=validated_data.storage_id,
                        original_sample_number=validated_data.original_sample_number,
                        source_id=validated_data.source_id,
                        location_id=validated_data.location_id,
                        appendix_note_id=validated_data.appendix_note_id,
                        comment_id=validated_data.comment_id,
                        has_photo=validated_data.has_photo,
                        is_identified=validated_data.is_identified,
                        has_antibiotic_activity=validated_data.has_antibiotic_activity,
                        has_genome=validated_data.has_genome,
                        has_biochemistry=validated_data.has_biochemistry,
                        seq_status=validated_data.seq_status,
                    )
                    logger.info(f"Created new sample after sequence reset (ID: {sample.id})")
            else:
                raise
        
        return Response({
            'id': sample.id,
            'message': f'Образец {sample.id} успешно создан',
            'sample': {
                'id': sample.id,
                'strain_id': sample.strain_id,
                'storage_id': sample.storage_id,
                'original_sample_number': sample.original_sample_number,
                'has_photo': sample.has_photo,
                'is_identified': sample.is_identified,
            }
        }, status=status.HTTP_201_CREATED)
    
    except ValidationError as e:
        return Response({
            'error': 'Ошибка валидации данных',
            'details': e.errors()
        }, status=status.HTTP_400_BAD_REQUEST)
    
    except Exception as e:
        logger.error(f"Unexpected error in create_sample: {e}")
        return Response({
            'error': f'Ошибка при создании образца: {e}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_sample(request, sample_id):
    """Получение детальной информации об образце"""
    try:
        sample = Sample.objects.select_related(
            'strain', 'storage', 'source', 'location', 'index_letter', 'appendix_note', 'comment'
        ).get(id=sample_id)
        
        return Response({
            'id': sample.id,
            'strain': {
                'id': sample.strain.id if sample.strain else None,
                'short_code': sample.strain.short_code if sample.strain else None,
                'identifier': sample.strain.identifier if sample.strain else None,
                'rrna_taxonomy': sample.strain.rrna_taxonomy if sample.strain else None,
            } if sample.strain else None,
            'storage': {
                'id': sample.storage.id if sample.storage else None,
                'box_id': sample.storage.box_id if sample.storage else None,
                'cell_id': sample.storage.cell_id if sample.storage else None,
            } if sample.storage else None,
            'source': {
                'id': sample.source.id if sample.source else None,
                'organism_name': sample.source.organism_name if sample.source else None,
                'source_type': sample.source.source_type if sample.source else None,
                'category': sample.source.category if sample.source else None,
            } if sample.source else None,
            'location': {
                'id': sample.location.id if sample.location else None,
                'name': sample.location.name if sample.location else None,
            } if sample.location else None,
            'index_letter': {
                'id': sample.index_letter.id if sample.index_letter else None,
                'letter_value': sample.index_letter.letter_value if sample.index_letter else None,
            } if sample.index_letter else None,
            'appendix_note': {
                'id': sample.appendix_note.id if sample.appendix_note else None,
                'text': sample.appendix_note.text if sample.appendix_note else None,
            } if sample.appendix_note else None,
            'comment': {
                'id': sample.comment.id if sample.comment else None,
                'text': sample.comment.text if sample.comment else None,
            } if sample.comment else None,
            'original_sample_number': sample.original_sample_number,
            'has_photo': sample.photos.exists(),
            'photos': [
                {
                    'id': photo.id,
                    'url': request.build_absolute_uri(photo.image.url),
                    'uploaded_at': photo.uploaded_at.isoformat()
                } for photo in sample.photos.all()
            ],
            'is_identified': sample.is_identified,
            'has_antibiotic_activity': sample.has_antibiotic_activity,
            'has_genome': sample.has_genome,
            'has_biochemistry': sample.has_biochemistry,
            'seq_status': sample.seq_status,
            'created_at': sample.created_at.isoformat() if sample.created_at else None,
            'updated_at': sample.updated_at.isoformat() if sample.updated_at else None,
        })
    
    except Sample.DoesNotExist:
        return Response({
            'error': f'Образец с ID {sample_id} не найден'
        }, status=status.HTTP_404_NOT_FOUND)
    
    except Exception as e:
        logger.error(f"Unexpected error in get_sample: {e}")
        return Response({
            'error': f'Ошибка при получении образца: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['PUT', 'PATCH'])
def update_sample(request, sample_id):
    """Обновление данных образца"""
    try:
        sample = Sample.objects.get(id=sample_id)
    except Sample.DoesNotExist:
        return Response({
            'error': f'Образец с ID {sample_id} не найден'
        }, status=status.HTTP_404_NOT_FOUND)
    
    try:
        # Динамически создаем схему валидации для обновления
        class UpdateSampleSchema(BaseModel):
            strain_id: Optional[int] = Field(None, ge=1, description="ID штамма")
            index_letter_id: Optional[int] = Field(None, ge=1, description="ID индексной буквы")
            storage_id: Optional[int] = Field(None, ge=1, description="ID хранилища")
            original_sample_number: Optional[str] = Field(None, max_length=100, description="Оригинальный номер образца")
            source_id: Optional[int] = Field(None, ge=1, description="ID источника")
            location_id: Optional[int] = Field(None, ge=1, description="ID местоположения")
            appendix_note_id: Optional[int] = Field(None, ge=1, description="ID примечания")
            comment_id: Optional[int] = Field(None, ge=1, description="ID комментария")
            has_photo: Optional[bool] = Field(None, description="Есть ли фото")
            is_identified: Optional[bool] = Field(None, description="Идентифицирован ли")
            has_antibiotic_activity: Optional[bool] = Field(None, description="Есть ли антибиотическая активность")
            has_genome: Optional[bool] = Field(None, description="Есть ли геном")
            has_biochemistry: Optional[bool] = Field(None, description="Есть ли биохимия")
            seq_status: Optional[bool] = Field(None, description="Статус секвенирования")
            
            @field_validator('original_sample_number')
            @classmethod
            def validate_sample_number(cls, v: Optional[str]) -> Optional[str]:
                if v is not None:
                    v = v.strip()
                    return v if v else None
                return v

        validated_data = UpdateSampleSchema.model_validate(request.data)
        
        # Проверяем существование связанных объектов если они указаны
        if validated_data.strain_id is not None:
            if not Strain.objects.filter(id=validated_data.strain_id).exists():
                return Response({
                    'error': f'Штамм с ID {validated_data.strain_id} не найден'
                }, status=status.HTTP_404_NOT_FOUND)
        
        if validated_data.storage_id is not None:
            if not Storage.objects.filter(id=validated_data.storage_id).exists():
                return Response({
                    'error': f'Хранилище с ID {validated_data.storage_id} не найдено'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Проверяем, не занята ли ячейка другим образцом
            existing_sample = Sample.objects.filter(
                storage_id=validated_data.storage_id
            ).exclude(id=sample_id).exclude(comment_id=2).first()
            
            if existing_sample:
                storage = Storage.objects.get(id=validated_data.storage_id)
                return Response({
                    'error': f'Ячейка {storage.box_id}-{storage.cell_id} уже занята образцом {existing_sample.id}'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # Аналогичные проверки для других связанных объектов
        if validated_data.source_id is not None and not Source.objects.filter(id=validated_data.source_id).exists():
            return Response({
                'error': f'Источник с ID {validated_data.source_id} не найден'
            }, status=status.HTTP_404_NOT_FOUND)
        
        if validated_data.location_id is not None and not Location.objects.filter(id=validated_data.location_id).exists():
            return Response({
                'error': f'Местоположение с ID {validated_data.location_id} не найдено'
            }, status=status.HTTP_404_NOT_FOUND)
        
        if validated_data.index_letter_id is not None and not IndexLetter.objects.filter(id=validated_data.index_letter_id).exists():
            return Response({
                'error': f'Индексная буква с ID {validated_data.index_letter_id} не найдена'
            }, status=status.HTTP_404_NOT_FOUND)
        
        if validated_data.appendix_note_id is not None and not AppendixNote.objects.filter(id=validated_data.appendix_note_id).exists():
            return Response({
                'error': f'Примечание с ID {validated_data.appendix_note_id} не найдено'
            }, status=status.HTTP_404_NOT_FOUND)
        
        if validated_data.comment_id is not None and not Comment.objects.filter(id=validated_data.comment_id).exists():
            return Response({
                'error': f'Комментарий с ID {validated_data.comment_id} не найден'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Обновление данных в транзакции
        with transaction.atomic():
            # Обновляем только переданные поля
            update_fields = []
            for field, value in validated_data.model_dump(exclude_unset=True).items():
                setattr(sample, field, value)
                update_fields.append(field)
            
            if update_fields:
                sample.save(update_fields=update_fields)
                logger.info(f"Updated sample {sample.id}, fields: {update_fields}")
            
            return Response({
                'id': sample.id,
                'message': f'Образец {sample.id} успешно обновлен',
                'updated_fields': update_fields
            })
    
    except ValidationError as e:
        return Response({
            'error': 'Ошибка валидации данных',
            'details': e.errors()
        }, status=status.HTTP_400_BAD_REQUEST)
    
    except Exception as e:
        logger.error(f"Unexpected error in update_sample: {e}")
        return Response({
            'error': f'Ошибка при обновлении образца: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['DELETE'])
def delete_sample(request, sample_id):
    """Удаление образца"""
    try:
        sample = Sample.objects.get(id=sample_id)
    except Sample.DoesNotExist:
        return Response({
            'error': f'Образец с ID {sample_id} не найден'
        }, status=status.HTTP_404_NOT_FOUND)
    
    try:
        with transaction.atomic():
            sample_info = f"образец {sample.id}"
            if sample.strain:
                sample_info += f" штамма {sample.strain.short_code}"
            if sample.storage:
                sample_info += f" из ячейки {sample.storage.box_id}-{sample.storage.cell_id}"
            
            sample.delete()
            logger.info(f"Deleted sample {sample_id}")
            
            return Response({
                'message': f'Образец {sample_id} успешно удален',
                'deleted_sample': sample_info
            })
    
    except Exception as e:
        logger.error(f"Unexpected error in delete_sample: {e}")
        return Response({
            'error': f'Ошибка при удалении образца: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_reference_data(request):
    """Получение справочных данных для форм"""
    try:
        # Получаем только активные/используемые данные
        # Увеличиваем лимит и добавляем поиск
        search_query = request.GET.get('search', '')
        
        strains_qs = Strain.objects.all()
        if search_query:
            strains_qs = strains_qs.filter(
                Q(short_code__icontains=search_query) |
                Q(identifier__icontains=search_query) |
                Q(rrna_taxonomy__icontains=search_query)
            )
        
        strains = strains_qs.order_by('short_code')[:500]  # Увеличили лимит
        sources = Source.objects.all().order_by('organism_name')[:100]
        locations = Location.objects.all().order_by('name')[:100]
        index_letters = IndexLetter.objects.all().order_by('letter_value')
        comments = Comment.objects.all().order_by('id')[:50]
        appendix_notes = AppendixNote.objects.all().order_by('id')[:50]
        
        # Получаем свободные ячейки хранения
        occupied_storage_ids = Sample.objects.filter(
            strain_id__isnull=False  # Только если есть штамм
        ).exclude(comment_id=2).values_list('storage_id', flat=True)
        
        free_storage = Storage.objects.exclude(
            id__in=occupied_storage_ids
        ).order_by('box_id', 'cell_id')[:200]
        
        return Response({
            'strains': [
                {
                    'id': strain.id,
                    'short_code': strain.short_code,
                    'identifier': strain.identifier,
                    'display_name': f"{strain.short_code} - {strain.identifier}"
                }
                for strain in strains
            ],
            'sources': [
                {
                    'id': source.id,
                    'organism_name': source.organism_name,
                    'source_type': source.source_type,
                    'category': source.category,
                    'display_name': f"{source.organism_name} ({source.source_type})"
                }
                for source in sources
            ],
            'locations': [
                {
                    'id': location.id,
                    'name': location.name
                }
                for location in locations
            ],
            'index_letters': [
                {
                    'id': letter.id,
                    'letter_value': letter.letter_value
                }
                for letter in index_letters
            ],
            'free_storage': [
                {
                    'id': storage.id,
                    'box_id': storage.box_id,
                    'cell_id': storage.cell_id,
                    'display_name': f"Бокс {storage.box_id}, ячейка {storage.cell_id}"
                }
                for storage in free_storage
            ],
            'comments': [
                {
                    'id': comment.id,
                    'text': comment.text[:100] + ('...' if len(comment.text) > 100 else '')
                }
                for comment in comments
            ],
            'appendix_notes': [
                {
                    'id': note.id,
                    'text': note.text[:100] + ('...' if len(note.text) > 100 else '')
                }
                for note in appendix_notes
            ]
        })
    
    except Exception as e:
        logger.error(f"Unexpected error in get_reference_data: {e}")
        return Response({
            'error': f'Ошибка при получении справочных данных: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_source_types(request):
    """Получение списка уникальных типов источников для выпадающего списка"""
    try:
        source_types = Source.objects.values_list('source_type', flat=True).distinct().order_by('source_type')
        
        return Response({
            'source_types': [
                {
                    'value': source_type,
                    'label': source_type
                }
                for source_type in source_types if source_type
            ]
        })
    
    except Exception as e:
        logger.error(f"Unexpected error in get_source_types: {e}")
        return Response({
            'error': f'Ошибка при получении типов источников: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_organism_names(request):
    """Получение списка уникальных названий организмов для выпадающего списка"""
    try:
        organism_names = Source.objects.values_list('organism_name', flat=True).distinct().order_by('organism_name')
        
        return Response({
            'organism_names': [
                {
                    'value': organism_name,
                    'label': organism_name
                }
                for organism_name in organism_names if organism_name
            ]
        })
    
    except Exception as e:
        logger.error(f"Unexpected error in get_organism_names: {e}")
        return Response({
            'error': f'Ошибка при получении названий организмов: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_boxes(request):
    """Получение списка боксов из обеих таблиц: StorageBox и Storage"""
    try:
        search_query = request.GET.get('search', '').strip()
        
        boxes = []
        
        # 1. Получаем боксы из таблицы StorageBox (новые боксы)
        storage_boxes_queryset = StorageBox.objects.all()
        
        if search_query:
            storage_boxes_queryset = storage_boxes_queryset.filter(
                Q(box_id__icontains=search_query) |
                Q(description__icontains=search_query)
            )
        
        for box in storage_boxes_queryset:
            # Вычисляем общее количество ячеек
            total_cells = box.rows * box.cols
            
            # Получаем количество занятых ячеек для этого бокса
            occupied_count = Sample.objects.filter(
                storage__box_id=box.box_id,
                strain_id__isnull=False
            ).exclude(comment_id=2).count()
            
            free_cells = total_cells - occupied_count
            
            display_id = box.box_id if str(box.box_id).startswith('Бокс') else f"Бокс {box.box_id}"
            
            boxes.append({
                'box_id': box.box_id,
                'total_cells': total_cells,
                'free_cells': free_cells,
                'rows': box.rows,
                'cols': box.cols,
                'description': box.description,
                'display_name': f"{display_id} ({free_cells} свободных из {total_cells})"
            })
        
        # 2. Получаем боксы из таблицы Storage (существующие боксы)
        # Исключаем боксы, которые уже есть в StorageBox
        existing_box_ids = set(storage_boxes_queryset.values_list('box_id', flat=True))
        
        with connection.cursor() as cursor:
            sql = """
                SELECT
                    s.box_id,
                    COUNT(*) as total_cells,
                    COUNT(*) - COUNT(occupied.id) as free_cells
                FROM collection_manager_storage s
                LEFT JOIN (
                    SELECT storage_id as id
                    FROM collection_manager_sample
                    WHERE storage_id IS NOT NULL
                    AND (comment_id IS NULL OR comment_id != 2)
                    AND strain_id IS NOT NULL
                ) occupied ON s.id = occupied.id
                WHERE s.box_id ILIKE %s
                GROUP BY s.box_id
                ORDER BY s.box_id
            """
            
            search_param = f"%{search_query}%" if search_query else "%"
            cursor.execute(sql, [search_param])
            
            for row in cursor.fetchall():
                box_id, total_cells, free_cells = row
                
                # Пропускаем боксы, которые уже есть в StorageBox
                if box_id not in existing_box_ids:
                    # Определяем размеры сетки из существующих ячеек
                    storage_cells = Storage.objects.filter(box_id=box_id)
                    max_row_letter = 'A'
                    max_col_num = 1
                    
                    for cell in storage_cells:
                        cell_id = cell.cell_id
                        if cell_id and len(cell_id) >= 2:
                            letter = cell_id[0]
                            try:
                                number = int(cell_id[1:])
                                if letter > max_row_letter:
                                    max_row_letter = letter
                                if number > max_col_num:
                                    max_col_num = number
                            except ValueError:
                                continue
                    
                    rows = ord(max_row_letter) - ord('A') + 1
                    cols = max_col_num
                    
                    # Пересчитываем total_cells на основе размеров сетки
                    calculated_total_cells = rows * cols
                    # Пересчитываем free_cells
                    occupied_count = total_cells - free_cells
                    calculated_free_cells = calculated_total_cells - occupied_count
                    
                    display_id = box_id if str(box_id).startswith('Бокс') else f"Бокс {box_id}"
                    
                    boxes.append({
                        'box_id': box_id,
                        'total_cells': calculated_total_cells,
                        'free_cells': calculated_free_cells,
                        'rows': rows,
                        'cols': cols,
                        'description': None,  # Нет описания для старых боксов
                        'display_name': f"{display_id} ({calculated_free_cells} свободных из {calculated_total_cells})"
                    })
        
        # Натуральная сортировка: если в ID есть число — сортируем по числу, иначе по строке
        def _box_sort_key(item):
            m = re.search(r"(\d+)$", str(item['box_id']))
            if m:
                return (0, int(m.group(1)))  # сначала числовые, по возрастанию
            return (1, str(item['box_id']))  # затем остальные лексикографически

        boxes.sort(key=_box_sort_key)
        
        return Response({'boxes': boxes})
    
    except Exception as e:
        logger.error(f"Unexpected error in get_boxes: {e}")
        return Response({
            'error': f'Ошибка при получении списка боксов: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_box_detail(request, box_id):
    """Получение детальной информации о боксе с сеткой ячеек"""
    try:
        # Сначала проверяем, есть ли бокс в новой таблице StorageBox
        storage_box = None
        try:
            storage_box = StorageBox.objects.get(box_id=box_id)
            rows = storage_box.rows
            cols = storage_box.cols
            description = storage_box.description
        except StorageBox.DoesNotExist:
            # Если нет в новой таблице, проверяем в старой таблице Storage
            storage_cells = Storage.objects.filter(box_id=box_id)
            if not storage_cells.exists():
                return Response({
                    'error': f'Бокс с ID {box_id} не найден'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Определяем размеры сетки из существующих ячеек
            # Парсим cell_id для определения максимальных размеров
            max_row_letter = 'A'
            max_col_num = 1
            
            for cell in storage_cells:
                cell_id = cell.cell_id
                if cell_id and len(cell_id) >= 2:
                    letter = cell_id[0]
                    number = int(cell_id[1:])
                    if letter > max_row_letter:
                        max_row_letter = letter
                    if number > max_col_num:
                        max_col_num = number
            
            rows = ord(max_row_letter) - ord('A') + 1
            cols = max_col_num
            description = None
        
        # Получаем все ячейки из таблицы Storage для этого бокса
        storage_cells = Storage.objects.filter(box_id=box_id)
        storage_cells_dict = {cell.cell_id: cell for cell in storage_cells}
        
        # Получаем информацию о занятых ячейках
        occupied_storage_ids = Sample.objects.filter(
            strain_id__isnull=False  # Только если есть штамм
        ).exclude(comment_id=2).values_list('storage_id', flat=True)
        
        # Создаем сетку ячеек
        cells_grid = []
        for row in range(rows):
            row_cells = []
            row_letter = chr(ord('A') + row)
            
            for col in range(1, cols + 1):
                cell_id = f"{row_letter}{col}"
                
                # Ищем ячейку в базе данных
                if cell_id in storage_cells_dict:
                    storage_cell = storage_cells_dict[cell_id]
                    is_occupied = storage_cell.id in occupied_storage_ids
                    
                    # Получаем информацию об образце, если ячейка занята
                    sample_info = None
                    if is_occupied:
                        samples = Sample.objects.filter(storage_id=storage_cell.id).exclude(comment_id=2).select_related('strain')
                        if samples.exists():
                            sample = samples.first()  # Берем первый образец
                            sample_info = {
                                'sample_id': sample.id,
                                'strain_id': sample.strain_id,
                                'strain_number': sample.strain.short_code if sample.strain else None,
                                'comment': sample.comment.text if sample.comment else None,
                                'total_samples': samples.count()  # Количество образцов в ячейке
                            }
                    
                    cell_data = {
                        'row': row + 1,
                        'col': col,
                        'cell_id': cell_id,
                        'storage_id': storage_cell.id,
                        'is_occupied': is_occupied,
                        'sample_info': sample_info
                    }
                else:
                    # Ячейка не существует в базе данных
                    cell_data = {
                        'row': row + 1,
                        'col': col,
                        'cell_id': cell_id,
                        'storage_id': None,
                        'is_occupied': False,
                        'sample_info': None
                    }
                
                row_cells.append(cell_data)
            cells_grid.append(row_cells)
        
        # Подсчитываем статистику
        total_cells = rows * cols
        occupied_cells = len([cell for row in cells_grid for cell in row if cell['is_occupied']])
        free_cells = total_cells - occupied_cells
        
        box_detail = {
            'box_id': box_id,
            'rows': rows,
            'cols': cols,
            'description': description,
            'total_cells': total_cells,
            'occupied_cells': occupied_cells,
            'free_cells': free_cells,
            'occupancy_percentage': round((occupied_cells / total_cells) * 100, 1) if total_cells > 0 else 0,
            'cells_grid': cells_grid
        }
        
        return Response(box_detail)
        
    except Exception as e:
        logger.error(f"Unexpected error in get_box_detail: {e}")
        return Response({
            'error': f'Ошибка при получении детальной информации о боксе: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_box_cells(request, box_id):
    """Получение списка свободных ячеек в указанном боксе"""
    try:
        search_query = request.GET.get('search', '').strip()
        
        # Получаем занятые storage_id (исключая свободные ячейки с comment_id=2)
        occupied_storage_ids = Sample.objects.filter(strain_id__isnull=False).exclude(comment_id=2).values_list('storage_id', flat=True)
        
        # Получаем свободные ячейки в указанном боксе
        cells_qs = Storage.objects.filter(box_id=box_id).exclude(
            id__in=occupied_storage_ids
        )
        
        if search_query:
            cells_qs = cells_qs.filter(
                Q(cell_id__icontains=search_query) |
                Q(box_id__icontains=search_query)
            )
        
        cells = cells_qs.order_by('cell_id')[:100]
        
        return Response({
            'box_id': box_id,
            'cells': [
                {
                    'id': cell.id,
                    'box_id': cell.box_id,
                    'cell_id': cell.cell_id,
                    'display_name': f"Ячейка {cell.cell_id}"
                }
                for cell in cells
            ]
        })
    
    except Exception as e:
        logger.error(f"Unexpected error in get_box_cells: {e}")
        return Response({
            'error': f'Ошибка при получении ячеек бокса {box_id}: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def bulk_delete_samples(request):
    """Массовое удаление образцов"""
    try:
        sample_ids = request.data.get('sample_ids', [])
        
        if not sample_ids:
            return Response({
                'error': 'Не указаны ID образцов для удаления'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not isinstance(sample_ids, list):
            return Response({
                'error': 'sample_ids должен быть списком'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Проверяем существование образцов
        existing_samples = Sample.objects.filter(id__in=sample_ids)
        existing_ids = list(existing_samples.values_list('id', flat=True))
        missing_ids = set(sample_ids) - set(existing_ids)
        
        if missing_ids:
            return Response({
                'error': f'Образцы с ID {list(missing_ids)} не найдены'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Собираем информацию об удаляемых образцах для логирования
        samples_info = []
        batch_id = generate_batch_id()
        
        for sample in existing_samples:
            info = f"ID {sample.id}"
            if sample.strain:
                info += f" (штамм {sample.strain.short_code})"
            if sample.storage:
                info += f" в ячейке {sample.storage.box_id}-{sample.storage.cell_id}"
            samples_info.append(info)
            
            # Логируем каждый удаляемый образец
            old_values = model_to_dict(sample)
            log_change(
                request=request,
                content_type='sample',
                object_id=sample.id,
                action='BULK_DELETE',
                old_values=old_values,
                comment=f"Массовое удаление {len(sample_ids)} образцов",
                batch_id=batch_id
            )
        
        # Выполняем массовое удаление
        with transaction.atomic():
            deleted_count = existing_samples.delete()[0]
            logger.info(f"Bulk deleted {deleted_count} samples: {sample_ids} (batch: {batch_id})")
        
        return Response({
            'message': f'Успешно удалено {deleted_count} образцов',
            'deleted_count': deleted_count,
            'deleted_samples': samples_info
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error in bulk_delete_samples: {e}")
        return Response({
            'error': f'Ошибка при массовом удалении: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def bulk_update_samples(request):
    """Массовое обновление образцов"""
    try:
        sample_ids = request.data.get('sample_ids', [])
        update_data = request.data.get('update_data', {})
        
        if not sample_ids:
            return Response({
                'error': 'Не указаны ID образцов для обновления'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not update_data:
            return Response({
                'error': 'Не указаны данные для обновления'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Проверяем существование образцов
        existing_samples = Sample.objects.filter(id__in=sample_ids)
        existing_ids = list(existing_samples.values_list('id', flat=True))
        missing_ids = set(sample_ids) - set(existing_ids)
        
        if missing_ids:
            return Response({
                'error': f'Образцы с ID {list(missing_ids)} не найдены'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Фильтруем только разрешенные поля для обновления
        allowed_fields = {
            'has_photo', 'is_identified', 'has_antibiotic_activity',
            'has_genome', 'has_biochemistry', 'seq_status'
        }
        
        filtered_update_data = {
            key: value for key, value in update_data.items() 
            if key in allowed_fields
        }
        
        if not filtered_update_data:
            return Response({
                'error': 'Нет допустимых полей для обновления'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Выполняем массовое обновление с логированием
        batch_id = generate_batch_id()
        
        with transaction.atomic():
            # Логируем каждый обновляемый образец до изменения
            for sample in existing_samples:
                old_values = model_to_dict(sample)
                
                # Применяем изменения к объекту для получения новых значений
                for field, value in filtered_update_data.items():
                    setattr(sample, field, value)
                
                new_values = model_to_dict(sample)
                
                log_change(
                    request=request,
                    content_type='sample',
                    object_id=sample.id,
                    action='BULK_UPDATE',
                    old_values=old_values,
                    new_values=new_values,
                    comment=f"Массовое обновление {len(sample_ids)} образцов: {list(filtered_update_data.keys())}",
                    batch_id=batch_id
                )
            
            # Выполняем массовое обновление
            updated_count = existing_samples.update(**filtered_update_data)
            logger.info(f"Bulk updated {updated_count} samples with data: {filtered_update_data} (batch: {batch_id})")
        
        return Response({
            'message': f'Успешно обновлено {updated_count} образцов',
            'updated_count': updated_count,
            'updated_fields': list(filtered_update_data.keys()),
            'updated_data': filtered_update_data
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error in bulk_update_samples: {e}")
        return Response({
            'error': f'Ошибка при массовом обновлении: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def bulk_delete_strains(request):
    """Массовое удаление штаммов"""
    try:
        strain_ids = request.data.get('strain_ids', [])
        force_delete = request.data.get('force_delete', False)
        
        if not strain_ids:
            return Response({
                'error': 'Не указаны ID штаммов для удаления'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Проверяем существование штаммов
        existing_strains = Strain.objects.filter(id__in=strain_ids)
        existing_ids = list(existing_strains.values_list('id', flat=True))
        missing_ids = set(strain_ids) - set(existing_ids)
        
        if missing_ids:
            return Response({
                'error': f'Штаммы с ID {list(missing_ids)} не найдены'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Проверяем связанные образцы
        strains_with_samples = []
        for strain in existing_strains:
            sample_count = Sample.objects.filter(strain_id=strain.id).count()
            if sample_count > 0:
                strains_with_samples.append({
                    'id': strain.id,
                    'short_code': strain.short_code,
                    'sample_count': sample_count
                })
        
        if strains_with_samples and not force_delete:
            return Response({
                'error': 'Некоторые штаммы имеют связанные образцы',
                'strains_with_samples': strains_with_samples,
                'message': 'Используйте force_delete=true для принудительного удаления'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Собираем информацию об удаляемых штаммах
        strains_info = []
        for strain in existing_strains:
            strains_info.append({
                'id': strain.id,
                'short_code': strain.short_code,
                'identifier': strain.identifier
            })
        
        # Выполняем массовое удаление
        with transaction.atomic():
            if force_delete:
                # Удаляем связанные образцы
                for strain in existing_strains:
                    Sample.objects.filter(strain_id=strain.id).delete()
            
            deleted_count = existing_strains.delete()[0]
            logger.info(f"Bulk deleted {deleted_count} strains: {strain_ids}")
        
        return Response({
            'message': f'Успешно удалено {deleted_count} штаммов',
            'deleted_count': deleted_count,
            'deleted_strains': strains_info,
            'force_delete': force_delete
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error in bulk_delete_strains: {e}")
        return Response({
            'error': f'Ошибка при массовом удалении штаммов: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def bulk_update_strains(request):
    """Массовое обновление штаммов"""
    try:
        strain_ids = request.data.get('strain_ids', [])
        update_data = request.data.get('update_data', {})
        
        if not strain_ids:
            return Response({
                'error': 'Не указаны ID штаммов для обновления'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not update_data:
            return Response({
                'error': 'Не указаны данные для обновления'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Проверяем существование штаммов
        existing_strains = Strain.objects.filter(id__in=strain_ids)
        existing_ids = list(existing_strains.values_list('id', flat=True))
        missing_ids = set(strain_ids) - set(existing_ids)
        
        if missing_ids:
            return Response({
                'error': f'Штаммы с ID {list(missing_ids)} не найдены'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Фильтруем только разрешенные поля для обновления штаммов
        allowed_fields = {
            'short_code', 'rrna_taxonomy', 'identifier', 'name_alt', 'rcam_collection_id'
        }
        
        filtered_update_data = {
            key: value for key, value in update_data.items() 
            if key in allowed_fields and value is not None
        }
        
        if not filtered_update_data:
            return Response({
                'error': 'Нет допустимых полей для обновления'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Выполняем массовое обновление с логированием
        batch_id = generate_batch_id()
        
        with transaction.atomic():
            # Логируем каждый обновляемый штамм до изменения
            for strain in existing_strains:
                old_values = model_to_dict(strain)
                
                # Применяем изменения к объекту для получения новых значений
                for field, value in filtered_update_data.items():
                    setattr(strain, field, value)
                
                new_values = model_to_dict(strain)
                
                log_change(
                    request=request,
                    content_type='strain',
                    object_id=strain.id,
                    action='BULK_UPDATE',
                    old_values=old_values,
                    new_values=new_values,
                    comment=f"Массовое обновление {len(strain_ids)} штаммов: {list(filtered_update_data.keys())}",
                    batch_id=batch_id
                )
            
            # Выполняем массовое обновление
            updated_count = existing_strains.update(**filtered_update_data)
            logger.info(f"Bulk updated {updated_count} strains with data: {filtered_update_data} (batch: {batch_id})")
        
        return Response({
            'message': f'Успешно обновлено {updated_count} штаммов',
            'updated_count': updated_count,
            'updated_fields': list(filtered_update_data.keys()),
            'updated_data': filtered_update_data
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error in bulk_update_strains: {e}")
        return Response({
            'error': f'Ошибка при массовом обновлении штаммов: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET', 'POST'])
def bulk_export_samples(request):
    """Экспорт образцов в CSV/JSON/Excel"""
    try:
        import csv, json
        from io import StringIO, BytesIO

        # Читаем параметры (GET или POST)
        if request.method == 'POST':
            sample_ids = request.data.get('sample_ids', '')
            format_type = request.data.get('format', 'csv').lower()
            fields = request.data.get('fields', '').split(',') if request.data.get('fields') else []
            include_related = request.data.get('include_related', 'true').lower() == 'true'
        else:
            sample_ids = request.GET.get('sample_ids', '')
            format_type = request.GET.get('format', 'csv').lower()
            fields = request.GET.get('fields', '').split(',') if request.GET.get('fields') else []
            include_related = request.GET.get('include_related', 'true').lower() == 'true'

        # Получаем выборку
        if sample_ids:
            sample_ids = [int(x) for x in sample_ids.split(',') if x.isdigit()]
            samples = Sample.objects.filter(id__in=sample_ids)
        else:
            samples = Sample.objects.all()
            
            # Добавляем простую текстовую фильтрацию (поиск)
            search_query = request.data.get('search', '').strip() if request.method == 'POST' else request.GET.get('search', '').strip()
            if search_query:
                from django.db.models import Q
                q = Q(original_sample_number__icontains=search_query) | Q(strain__short_code__icontains=search_query) | Q(strain__identifier__icontains=search_query)
                samples = samples.filter(q)

        if include_related:
            samples = samples.select_related('strain', 'storage', 'source', 'location')

        # Доступные поля
        available_fields = {
            'id': 'ID',
            'strain_short_code': 'Код штамма',
            'strain_identifier': 'Идентификатор штамма',
            'original_sample_number': 'Номер образца',
            'has_photo': 'Есть фото',
            'is_identified': 'Идентифицирован',
            'has_antibiotic_activity': 'АБ активность',
            'has_genome': 'Есть геном',
            'has_biochemistry': 'Есть биохимия',
            'seq_status': 'Секвенирован',
            'source_organism': 'Организм источника',
            'source_type': 'Тип источника',
            'location_name': 'Локация',
            'storage_cell': 'Ячейка хранения',
            'created_at': 'Дата создания',
            'updated_at': 'Дата обновления',
        }

        if not fields or not any(f.strip() for f in fields):
            export_fields = list(available_fields.keys())
        else:
            export_fields = [f.strip() for f in fields if f.strip() in available_fields]

        # Собираем данные
        data = []
        for s in samples:
            row = {}
            for f in export_fields:
                if f == 'id':
                    row[available_fields[f]] = s.id
                elif f == 'strain_short_code':
                    row[available_fields[f]] = s.strain.short_code if s.strain else ''
                elif f == 'strain_identifier':
                    row[available_fields[f]] = s.strain.identifier if s.strain else ''
                elif f == 'original_sample_number':
                    row[available_fields[f]] = s.original_sample_number or ''
                elif f == 'has_photo':
                    row[available_fields[f]] = 'Да' if s.has_photo else 'Нет'
                elif f == 'is_identified':
                    row[available_fields[f]] = 'Да' if s.is_identified else 'Нет'
                elif f == 'has_antibiotic_activity':
                    row[available_fields[f]] = 'Да' if s.has_antibiotic_activity else 'Нет'
                elif f == 'has_genome':
                    row[available_fields[f]] = 'Да' if s.has_genome else 'Нет'
                elif f == 'has_biochemistry':
                    row[available_fields[f]] = 'Да' if s.has_biochemistry else 'Нет'
                elif f == 'seq_status':
                    row[available_fields[f]] = 'Да' if s.seq_status else 'Нет'
                elif f == 'source_organism':
                    row[available_fields[f]] = s.source.organism_name if s.source else ''
                elif f == 'source_type':
                    row[available_fields[f]] = s.source.source_type if s.source else ''
                elif f == 'location_name':
                    row[available_fields[f]] = s.location.name if s.location else ''
                elif f == 'storage_cell':
                    row[available_fields[f]] = f"{s.storage.box_id}:{s.storage.cell_id}" if s.storage else ''
                elif f == 'created_at':
                    row[available_fields[f]] = s.created_at.strftime('%Y-%m-%d %H:%M:%S') if s.created_at else ''
                elif f == 'updated_at':
                    row[available_fields[f]] = s.updated_at.strftime('%Y-%m-%d %H:%M:%S') if s.updated_at else ''
            data.append(row)

        # Возврат в нужном формате
        from django.http import HttpResponse
        if format_type == 'json':
            return HttpResponse(json.dumps(data, ensure_ascii=False, indent=2), content_type='application/json; charset=utf-8')
        elif format_type == 'excel' or format_type == 'xlsx':
            try:
                import openpyxl
                from openpyxl import Workbook

                wb = Workbook()
                ws = wb.active
                if data:
                    ws.append(list(data[0].keys()))
                    for row in data:
                        ws.append(list(row.values()))
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="samples_export.xlsx"'
                return response
            except ImportError:
                return Response({'error':'Excel экспорт недоступен'}, status=500)
        else:
            # CSV
            output = StringIO()
            writer = csv.writer(output)
            if data:
                writer.writerow(list(data[0].keys()))
                for row in data:
                    writer.writerow(list(row.values()))
            response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = 'attachment; filename="samples_export.csv"'
            return response

    except Exception as e:
        logger.error(f"Error in bulk_export_samples: {e}")
        return Response({
            'error': f'Ошибка при экспорте: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET', 'POST'])
def bulk_export_strains(request):
    """Экспорт штаммов в CSV/JSON/Excel"""
    try:
        import csv
        import json
        from io import StringIO, BytesIO
        
        # Получаем параметры (поддерживаем GET и POST)
        if request.method == 'POST':
            strain_ids = request.data.get('strain_ids', '')
            format_type = request.data.get('format', 'csv').lower()
            fields = request.data.get('fields', '').split(',') if request.data.get('fields') else []
            include_related = request.data.get('include_related', 'true').lower() == 'true'
        else:
            strain_ids = request.GET.get('strain_ids', '')
            format_type = request.GET.get('format', 'csv').lower()
            fields = request.GET.get('fields', '').split(',') if request.GET.get('fields') else []
            include_related = request.GET.get('include_related', 'true').lower() == 'true'
        
        # Получаем штаммы
        if strain_ids:
            strain_ids = [int(x) for x in strain_ids.split(',') if x.isdigit()]
            strains = Strain.objects.filter(id__in=strain_ids)
        else:
            # Применяем те же фильтры, что и в list_strains
            strains = Strain.objects.all()
            
            # Фильтрация (копируем логику из list_strains)
            if request.method == 'POST':
                search_query = request.data.get('search', '').strip()
            else:
                search_query = request.GET.get('search', '').strip()
            
            if search_query:
                from django.db.models import Q
                search_q = Q()
                search_q |= Q(short_code__icontains=search_query)
                search_q |= Q(identifier__icontains=search_query)
                search_q |= Q(rrna_taxonomy__icontains=search_query)
                search_q |= Q(name_alt__icontains=search_query)
                search_q |= Q(rcam_collection_id__icontains=search_query)
                strains = strains.filter(search_q)
        
        strains = strains.order_by('id')
        
        # Определяем поля для экспорта
        available_fields = {
            'id': 'ID',
            'short_code': 'Короткий код',
            'identifier': 'Идентификатор',
            'rrna_taxonomy': 'rRNA таксономия',
            'name_alt': 'Альтернативное название',
            'rcam_collection_id': 'RCAM ID',
            'created_at': 'Дата создания',
            'updated_at': 'Дата обновления'
        }
        
        # Если поля не указаны, экспортируем все
        if not fields or not any(f.strip() for f in fields):
            export_fields = list(available_fields.keys())
        else:
            export_fields = [f.strip() for f in fields if f.strip() in available_fields]
        
        if not export_fields:
            return Response({
                'error': 'Не указаны корректные поля для экспорта'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Подготавливаем данные
        data = []
        for strain in strains:
            row = {}
            for field in export_fields:
                if field == 'id':
                    row[available_fields[field]] = strain.id
                elif field == 'short_code':
                    row[available_fields[field]] = strain.short_code or ''
                elif field == 'identifier':
                    row[available_fields[field]] = strain.identifier or ''
                elif field == 'rrna_taxonomy':
                    row[available_fields[field]] = strain.rrna_taxonomy or ''
                elif field == 'name_alt':
                    row[available_fields[field]] = strain.name_alt or ''
                elif field == 'rcam_collection_id':
                    row[available_fields[field]] = strain.rcam_collection_id or ''
                elif field == 'created_at':
                    row[available_fields[field]] = strain.created_at.strftime('%Y-%m-%d %H:%M:%S') if strain.created_at else ''
                elif field == 'updated_at':
                    row[available_fields[field]] = strain.updated_at.strftime('%Y-%m-%d %H:%M:%S') if strain.updated_at else ''
            data.append(row)
        
        # Экспорт в зависимости от формата
        if format_type == 'json':
            from django.http import HttpResponse
            response = HttpResponse(
                json.dumps(data, ensure_ascii=False, indent=2),
                content_type='application/json; charset=utf-8'
            )
            response['Content-Disposition'] = 'attachment; filename="strains_export.json"'
            return response
            
        elif format_type == 'excel':
            try:
                import openpyxl
                from openpyxl import Workbook
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Штаммы"
                
                # Заголовки
                if data:
                    headers = list(data[0].keys())
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)
                    
                    # Данные
                    for row_idx, row_data in enumerate(data, 2):
                        for col_idx, value in enumerate(row_data.values(), 1):
                            ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Сохраняем в BytesIO
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                from django.http import HttpResponse
                response = HttpResponse(
                    output.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename="strains_export.xlsx"'
                return response
                
            except ImportError:
                return Response({
                    'error': 'Excel экспорт недоступен. Установите openpyxl.'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        else:  # CSV по умолчанию
            output = StringIO()
            writer = csv.writer(output)

            # Заголовки и данные
            if data:
                headers = list(data[0].keys())
                writer.writerow(headers)
                for row_data in data:
                    writer.writerow(list(row_data.values()))

            from django.http import HttpResponse
            response = HttpResponse(
                output.getvalue(),
                content_type='text/csv; charset=utf-8'
            )
            response['Content-Disposition'] = 'attachment; filename="strains_export.csv"'
            return response
        
    except Exception as e:
        logger.error(f"Error in bulk_export_strains: {e}")
        return Response({
            'error': f'Ошибка при экспорте: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR) 


@api_view(['GET'])
def api_health(request):
    """Простой health-check endpoint для Docker"""
    return Response({'status': 'healthy'})


@api_view(['GET'])
def list_storage_summary(request):
    """Краткая информация о хранилищах без деталей ячеек (для быстрой загрузки)"""
    from django.db import connection
    
    # ИСПРАВЛЕННЫЙ SQL ЗАПРОС - правильная логика для NULL comment_id
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                s.box_id,
                COUNT(*) as total_cells,
                COUNT(CASE 
                    WHEN sam.strain_id IS NOT NULL AND (sam.comment_id IS NULL OR sam.comment_id != 2)
                    THEN 1 
                    ELSE NULL 
                END) as occupied_cells
            FROM collection_manager_storage s
            LEFT JOIN collection_manager_sample sam ON s.id = sam.storage_id
            GROUP BY s.box_id
            ORDER BY s.box_id
        """)
        
        boxes = []
        total_boxes = 0
        total_cells = 0
        occupied_cells = 0
        
        for row in cursor.fetchall():
            box_id, total, occupied = row
            boxes.append({
                'box_id': box_id,
                'occupied': occupied,
                'total': total
            })
            total_boxes += 1
            total_cells += total
            occupied_cells += occupied
    
    return Response({
        'boxes': boxes,
        'total_boxes': total_boxes,
        'total_cells': total_cells,
        'occupied_cells': occupied_cells
    })


@api_view(['GET'])
def get_box_details(request, box_id):
    """Детальная информация о ячейках конкретного бокса"""
    from django.db import connection
    
    # ИСПРАВЛЕННЫЙ SQL ЗАПРОС - правильная логика для NULL comment_id
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                s.id as storage_id,
                s.cell_id,
                sam.id as sample_id,
                st.short_code as strain_code,
                CASE 
                    WHEN sam.strain_id IS NOT NULL AND (sam.comment_id IS NULL OR sam.comment_id != 2)
                    THEN true 
                    ELSE false 
                END as occupied,
                CASE 
                    WHEN sam.id IS NULL OR sam.comment_id = 2 
                    THEN true 
                    ELSE false 
                END as is_free_cell
            FROM collection_manager_storage s
            LEFT JOIN collection_manager_sample sam ON s.id = sam.storage_id
            LEFT JOIN collection_manager_strain st ON sam.strain_id = st.id
            WHERE s.box_id = %s
            ORDER BY s.cell_id
        """, [box_id])
        
        cells = []
        occupied_count = 0
        
        for row in cursor.fetchall():
            storage_id, cell_id, sample_id, strain_code, occupied, is_free_cell = row
            
            cell_info = {
                'cell_id': cell_id,
                'storage_id': storage_id,
                'occupied': occupied,
                'sample_id': sample_id,
                'strain_code': strain_code,
                'is_free_cell': is_free_cell
            }
            
            cells.append(cell_info)
            if occupied:
                occupied_count += 1
    
    return Response({
        'box_id': box_id,
        'cells': cells,
        'total': len(cells),
        'occupied': occupied_count
    })


@api_view(['GET'])
def analytics_data(request):
    """Оптимизированный endpoint для аналитики - возвращает только агрегированные данные"""
    try:
        with connection.cursor() as cursor:
            # 1. Основная статистика - используем тот же подход что и в list_storage_summary
            cursor.execute("""
                SELECT 
                    (SELECT COUNT(*) FROM collection_manager_strain) as total_strains,
                    (SELECT COUNT(*) FROM collection_manager_sample) as total_samples
            """)
            basic_counts = cursor.fetchone()
            
            # Получаем правильное количество ячеек хранения (ТОЧНО как в list_storage_summary)
            cursor.execute("""
                SELECT 
                    s.box_id,
                    COUNT(*) as total_cells
                FROM collection_manager_storage s
                LEFT JOIN collection_manager_sample sam ON s.id = sam.storage_id
                GROUP BY s.box_id
            """)
            storage_rows = cursor.fetchall()
            total_storage_cells = sum(row[1] for row in storage_rows)
            
            counts = (basic_counts[0], basic_counts[1], total_storage_cells)
            
            # 2. Распределение по типам источников (агрегированно)
            cursor.execute("""
                SELECT 
                    src.source_type,
                    COUNT(sam.id) as count
                FROM collection_manager_sample sam
                LEFT JOIN collection_manager_source src ON sam.source_id = src.id
                WHERE src.source_type IS NOT NULL
                GROUP BY src.source_type
                ORDER BY count DESC
            """)
            source_distribution = {row[0]: row[1] for row in cursor.fetchall()}
            
            # 3. Топ 10 штаммов по количеству образцов
            cursor.execute("""
                SELECT 
                    st.short_code,
                    COUNT(sam.id) as count
                FROM collection_manager_sample sam
                LEFT JOIN collection_manager_strain st ON sam.strain_id = st.id
                WHERE st.short_code IS NOT NULL
                GROUP BY st.short_code
                ORDER BY count DESC
                LIMIT 10
            """)
            strain_distribution = {row[0]: row[1] for row in cursor.fetchall()}
            
            # 4. Статистика характеристик образцов
            cursor.execute("""
                SELECT 
                    SUM(CASE WHEN has_photo = true THEN 1 ELSE 0 END) as with_photo,
                    SUM(CASE WHEN is_identified = true THEN 1 ELSE 0 END) as identified,
                    SUM(CASE WHEN has_antibiotic_activity = true THEN 1 ELSE 0 END) as with_antibiotic_activity,
                    SUM(CASE WHEN has_genome = true THEN 1 ELSE 0 END) as with_genome,
                    SUM(CASE WHEN has_biochemistry = true THEN 1 ELSE 0 END) as with_biochemistry,
                    SUM(CASE WHEN seq_status = true THEN 1 ELSE 0 END) as sequenced
                FROM collection_manager_sample
            """)
            characteristics = cursor.fetchone()
            
            # 5. Утилизация хранилища (ИСПРАВЛЕННО - правильная логика как в list_storage_summary)
            cursor.execute("""
                SELECT 
                    SUM(CASE 
                        WHEN sam.strain_id IS NOT NULL AND (sam.comment_id IS NULL OR sam.comment_id != 2) 
                        THEN 1 ELSE 0 
                    END) as occupied,
                    SUM(CASE 
                        WHEN sam.strain_id IS NULL OR sam.comment_id = 2 
                        THEN 1 ELSE 0 
                    END) as free,
                    COUNT(*) as total
                FROM collection_manager_storage st
                LEFT JOIN collection_manager_sample sam ON st.id = sam.storage_id
            """)
            storage_stats = cursor.fetchone()
            
            # 6. Месячные тренды (последние 12 месяцев)
            cursor.execute("""
                SELECT 
                    DATE_TRUNC('month', created_at) as month,
                    COUNT(*) as count
                FROM collection_manager_sample
                WHERE created_at >= CURRENT_DATE - INTERVAL '12 months'
                GROUP BY DATE_TRUNC('month', created_at)
                ORDER BY month
            """)
            monthly_data = cursor.fetchall()
            
            # Формируем полные данные за 12 месяцев (включая месяцы с 0 образцов)
            from datetime import datetime, timedelta
            import calendar
            
            monthly_trends = []
            now = datetime.now()
            
            # Создаем словарь из данных БД
            db_months = {row[0].strftime('%Y-%m'): row[1] for row in monthly_data if row[0]}
            
            for i in range(11, -1, -1):
                date = datetime(now.year, now.month, 1) - timedelta(days=32*i)
                date = date.replace(day=1)  # Первое число месяца
                month_key = date.strftime('%Y-%m')
                month_name = f"{calendar.month_abbr[date.month]} {date.year}"
                count = db_months.get(month_key, 0)
                monthly_trends.append({
                    'month': month_name,
                    'count': count
                })
        
        # Формируем ответ
        analytics_data = {
            'totalSamples': counts[1],
            'totalStrains': counts[0], 
            'totalStorage': counts[2],
            'sourceTypeDistribution': source_distribution,
            'strainDistribution': strain_distribution,
            'monthlyTrends': monthly_trends,
            'characteristicsStats': {
                'has_photo': characteristics[0] or 0,
                'is_identified': characteristics[1] or 0,
                'has_antibiotic_activity': characteristics[2] or 0,
                'has_genome': characteristics[3] or 0,
                'has_biochemistry': characteristics[4] or 0,
                'seq_status': characteristics[5] or 0,
            },
            'storageUtilization': {
                'occupied': storage_stats[0] or 0,
                'free': storage_stats[1] or 0,
                'total': storage_stats[2] or 0
            }
        }
        
        return Response(analytics_data)
        
    except Exception as e:
        logger.error(f"Ошибка в analytics_data: {e}")
        return Response(
            {'error': f'Ошибка получения данных аналитики: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
def create_box(request):
    """Создание нового бокса с автоматической генерацией номера и ячеек"""
    try:
        class CreateBoxSchema(BaseModel):
            rows: int = Field(gt=0, le=99, description="Количество рядов")
            cols: int = Field(gt=0, le=99, description="Количество колонок")
            description: Optional[str] = Field(None, description="Описание бокса")

        try:
            data = CreateBoxSchema.model_validate(request.data)
        except ValidationError as e:
            return Response({'errors': e.errors()}, status=status.HTTP_400_BAD_REQUEST)

        from .models import StorageBox, Storage
        from django.db import transaction
        
        # Генерируем следующий номер бокса автоматически.
        # Нужно учесть два возможных формата существующих ID:
        #   1. «Бокс N»  – новый формат
        #   2. «N»       – старый формат (только число)

        # Получаем все box_id из обеих таблиц и вычисляем максимальный номер
        existing_ids = list(StorageBox.objects.values_list('box_id', flat=True))
        existing_ids += list(Storage.objects.values_list('box_id', flat=True).distinct())

        max_number = 0
        for bid in existing_ids:
            # Срезаем префикс «Бокс » при наличии
            if bid.startswith('Бокс '):
                bid = bid[5:].strip()

            try:
                num = int(bid)
                if num > max_number:
                    max_number = num
            except ValueError:
                continue

        next_number = max_number + 1
        
        # Новый ID бокса (без префикса «Бокс » — храним только число)
        box_id = str(next_number)

        # Проверяем уникальность и в StorageBox, и в Storage
        while (
            StorageBox.objects.filter(box_id=box_id).exists() or
            Storage.objects.filter(box_id=box_id).exists()
        ):
            next_number += 1
            box_id = str(next_number)

        with transaction.atomic():
            # Создаем бокс с автогенерированным ID
            box = StorageBox.objects.create(
                box_id=box_id,
                rows=data.rows,
                cols=data.cols,
                description=data.description
            )

            # Генерируем ячейки
            cells_to_create = []
            for row_idx in range(1, data.rows + 1):
                # Преобразуем номер ряда в букву (A=1, B=2...). Поддерживаются до Z (26)
                if row_idx <= 26:
                    row_letter = chr(64 + row_idx)  # 65->A
                else:
                    # Для >26 используем двойные буквы (AA, AB...)
                    first = chr(64 + ((row_idx - 1) // 26))
                    second = chr(65 + ((row_idx - 1) % 26))
                    row_letter = first + second

                for col_idx in range(1, data.cols + 1):
                    cell_num = str(col_idx)
                    cell_id = f"{row_letter}{cell_num}"
                    cells_to_create.append(Storage(box_id=box_id, cell_id=cell_id))

            # Используем bulk_create с ignore_conflicts для избежания дублирования
            Storage.objects.bulk_create(cells_to_create, batch_size=1000, ignore_conflicts=True)

            logger.info(f"Created box {box_id} with {len(cells_to_create)} cells")

            return Response({
                'message': f'Бокс "{box_id}" успешно создан',
                'box': {
                    'box_id': box.box_id,
                    'rows': box.rows,
                    'cols': box.cols,
                    'description': box.description,
                    'cells_created': len(cells_to_create),
                    'generated_id': True
                }
            }, status=status.HTTP_201_CREATED)

    except Exception as e:
        logger.error(f"Error in create_box: {e}")
        return Response({'error': f'Ошибка при создании бокса: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_box(request, box_id):
    """Получение детальной информации о боксе"""
    try:
        from .models import StorageBox, Storage
        
        # Получаем бокс
        try:
            box = StorageBox.objects.get(box_id=box_id)
        except StorageBox.DoesNotExist:
            return Response({
                'error': f'Бокс с ID {box_id} не найден'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Получаем статистику ячеек
        total_cells = Storage.objects.filter(box_id=box_id).count()
        
        # Получаем занятые ячейки (исключая помеченные как свободные)
        occupied_cells = Sample.objects.filter(
            storage__box_id=box_id,
            strain_id__isnull=False
        ).exclude(comment_id=2).count()
        
        # Получаем помеченные как свободные ячейки
        free_marked_cells = Sample.objects.filter(
            storage__box_id=box_id,
            comment_id=2
        ).count()
        
        empty_cells = total_cells - occupied_cells - free_marked_cells
        
        return Response({
            'box_id': box.box_id,
            'rows': box.rows,
            'cols': box.cols,
            'description': box.description,
            'created_at': box.created_at.isoformat() if box.created_at else None,
            'statistics': {
                'total_cells': total_cells,
                'occupied_cells': occupied_cells,
                'free_marked_cells': free_marked_cells,
                'empty_cells': empty_cells,
                'occupancy_percentage': round((occupied_cells / total_cells * 100) if total_cells > 0 else 0, 1)
            }
        })
    
    except Exception as e:
        logger.error(f"Error in get_box: {e}")
        return Response({
            'error': f'Ошибка при получении информации о боксе: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['PUT', 'PATCH'])
def update_box(request, box_id):
    """Обновление информации о боксе"""
    try:
        class UpdateBoxSchema(BaseModel):
            description: Optional[str] = Field(None, description="Описание бокса")
            
            @field_validator('description')
            @classmethod
            def validate_description(cls, v: Optional[str]) -> Optional[str]:
                if v is not None:
                    v = v.strip()
                    return v if v else None
                return v
        
        from .models import StorageBox
        
        # Проверяем существование бокса
        try:
            box = StorageBox.objects.get(box_id=box_id)
        except StorageBox.DoesNotExist:
            return Response({
                'error': f'Бокс с ID {box_id} не найден'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Валидация входных данных
        try:
            validated_data = UpdateBoxSchema.model_validate(request.data)
        except ValidationError as e:
            return Response({'errors': e.errors()}, status=status.HTTP_400_BAD_REQUEST)
        
        # Обновление бокса
        with transaction.atomic():
            if validated_data.description is not None:
                box.description = validated_data.description
            box.save()
            
            # Логируем изменение
            log_change(
                request=request,
                content_type='StorageBox',
                object_id=box.id,
                action='UPDATE',
                new_values={'description': validated_data.description},
                comment='Box description updated'
            )
        
        return Response({
            'message': 'Бокс успешно обновлен',
            'box': {
                'box_id': box.box_id,
                'rows': box.rows,
                'cols': box.cols,
                'description': box.description,
                'created_at': box.created_at.isoformat() if box.created_at else None
            }
        })
    
    except Exception as e:
        logger.error(f"Error in update_box: {e}")
        return Response({
            'error': f'Ошибка при обновлении бокса: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['DELETE'])
def delete_box(request, box_id):
    """Удаление бокса с проверкой занятых ячеек"""
    try:
        from .models import StorageBox, Storage
        
        # Проверяем существование бокса
        try:
            box = StorageBox.objects.get(box_id=box_id)
        except StorageBox.DoesNotExist:
            return Response({
                'error': f'Бокс с ID {box_id} не найден'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Проверяем наличие занятых ячеек
        occupied_cells_count = Sample.objects.filter(
            storage__box_id=box_id,
            strain_id__isnull=False
        ).exclude(comment_id=2).count()
        
        force_delete = request.GET.get('force', '').lower() == 'true'
        
        if occupied_cells_count > 0 and not force_delete:
            return Response({
                'error': f'Бокс содержит {occupied_cells_count} занятых ячеек. Используйте параметр ?force=true для принудительного удаления',
                'occupied_cells': occupied_cells_count,
                'can_force_delete': True
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Получаем статистику перед удалением
        total_cells = Storage.objects.filter(box_id=box_id).count()
        
        # Удаляем бокс и все связанные ячейки
        with transaction.atomic():
            # Сначала освобождаем все образцы в ячейках
            if occupied_cells_count > 0:
                Sample.objects.filter(storage__box_id=box_id).update(storage=None)
            
            # Удаляем все ячейки бокса
            deleted_cells = Storage.objects.filter(box_id=box_id).delete()
            
            # Сохраняем ID до удаления для логирования
            box_pk = box.id
            
            # Удаляем сам бокс
            box.delete()
            
            # Логируем удаление (используем сохранённый PK, т.к. после delete() он становится None)
            log_change(
                request=request,
                content_type='StorageBox',
                object_id=box_pk,
                action='DELETE',
                old_values={
                    'box_id': box_id,
                    'total_cells_deleted': total_cells,
                    'occupied_cells_freed': occupied_cells_count,
                    'force_delete': force_delete
                },
                comment='Box deleted with all cells'
            )
        
        return Response({
            'message': f'Бокс {box_id} успешно удален',
            'statistics': {
                'cells_deleted': total_cells,
                'samples_freed': occupied_cells_count,
                'force_delete_used': force_delete
            }
        })
    
    except Exception as e:
        logger.error(f"Error in delete_box: {e}")
        return Response({
            'error': f'Ошибка при удалении бокса: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['PUT'])
def assign_cell(request, box_id, cell_id):
    """Размещение образца в ячейке"""
    try:
        class AssignCellSchema(BaseModel):
            sample_id: int = Field(gt=0, description="ID образца для размещения")
            
        from .models import Storage
        
        # Валидация входных данных
        try:
            validated_data = AssignCellSchema.model_validate(request.data)
        except ValidationError as e:
            return Response({'errors': e.errors()}, status=status.HTTP_400_BAD_REQUEST)
        
        # Проверяем существование ячейки
        try:
            storage_cell = Storage.objects.get(box_id=box_id, cell_id=cell_id)
        except Storage.DoesNotExist:
            return Response({
                'error': f'Ячейка {cell_id} в боксе {box_id} не найдена'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Проверяем существование образца
        try:
            sample = Sample.objects.get(id=validated_data.sample_id)
        except Sample.DoesNotExist:
            return Response({
                'error': f'Образец с ID {validated_data.sample_id} не найден'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Проверяем, что ячейка свободна
        existing_sample = Sample.objects.filter(storage=storage_cell).exclude(comment_id=2).first()
        if existing_sample:
            return Response({
                'error': f'Ячейка {cell_id} уже занята образцом ID {existing_sample.id}',
                'occupied_by': {
                    'sample_id': existing_sample.id,
                    'strain_code': existing_sample.strain.short_code if existing_sample.strain else None
                }
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Проверяем, что образец еще не размещен в другой ячейке
        if sample.storage and sample.comment_id != 2:
            return Response({
                'error': f'Образец уже размещен в ячейке {sample.storage.cell_id} бокса {sample.storage.box_id}',
                'current_location': {
                    'box_id': sample.storage.box_id,
                    'cell_id': sample.storage.cell_id
                }
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Размещаем образец
        with transaction.atomic():
            sample.storage = storage_cell
            sample.comment_id = None  # Убираем пометку "свободная ячейка" если была
            sample.save()
            
            # Логируем размещение
            log_change(
                request=request,
                content_type='Sample',
                object_id=sample.id,
                action='ASSIGN_CELL',
                new_values={
                    'box_id': box_id,
                    'cell_id': cell_id,
                    'storage_id': storage_cell.id
                },
                comment='Sample assigned to cell'
            )
        
        return Response({
            'message': f'Образец успешно размещен в ячейке {cell_id}',
            'assignment': {
                'sample_id': sample.id,
                'box_id': box_id,
                'cell_id': cell_id,
                'strain_code': sample.strain.short_code if sample.strain else None
            }
        })
    
    except Exception as e:
        logger.error(f"Error in assign_cell: {e}")
        return Response({
            'error': f'Ошибка при размещении образца: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['DELETE'])
def clear_cell(request, box_id, cell_id):
    """Освобождение ячейки"""
    try:
        from .models import Storage
        
        # Проверяем существование ячейки
        try:
            storage_cell = Storage.objects.get(box_id=box_id, cell_id=cell_id)
        except Storage.DoesNotExist:
            return Response({
                'error': f'Ячейка {cell_id} в боксе {box_id} не найдена'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Находим образец в ячейке
        sample = Sample.objects.filter(storage=storage_cell).exclude(comment_id=2).first()
        if not sample:
            return Response({
                'error': f'Ячейка {cell_id} уже свободна'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Освобождаем ячейку
        with transaction.atomic():
            sample_info = {
                'sample_id': sample.id,
                'strain_code': sample.strain.short_code if sample.strain else None
            }
            
            sample.storage = None
            sample.save()
            
            # Логируем освобождение
            log_change(
                request=request,
                content_type='Sample',
                object_id=sample.id,
                action='CLEAR_CELL',
                old_values={
                    'box_id': box_id,
                    'cell_id': cell_id,
                    'freed_sample_id': sample.id
                },
                comment='Cell cleared, sample removed'
            )
        
        return Response({
            'message': f'Ячейка {cell_id} успешно освобождена',
            'freed_sample': sample_info
        })
    
    except Exception as e:
        logger.error(f"Error in clear_cell: {e}")
        return Response({
            'error': f'Ошибка при освобождении ячейки: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def bulk_assign_cells(request, box_id):
    """Массовое размещение образцов в ячейках"""
    try:
        class BulkAssignSchema(BaseModel):
            assignments: list = Field(description="Список назначений")
            
            class Assignment(BaseModel):
                cell_id: str = Field(description="ID ячейки")
                sample_id: int = Field(gt=0, description="ID образца")
        
        from .models import Storage
        
        # Валидация входных данных
        try:
            validated_data = BulkAssignSchema.model_validate(request.data)
        except ValidationError as e:
            return Response({'errors': e.errors()}, status=status.HTTP_400_BAD_REQUEST)
        
        if not validated_data.assignments:
            return Response({
                'error': 'Список назначений не может быть пустым'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Валидируем каждое назначение
        assignments = []
        for i, assignment_data in enumerate(validated_data.assignments):
            try:
                assignment = BulkAssignSchema.Assignment.model_validate(assignment_data)
                assignments.append(assignment)
            except ValidationError as e:
                return Response({
                    'error': f'Ошибка в назначении #{i+1}: {e.errors()}'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # Проверяем все ячейки и образцы
        errors = []
        success_assignments = []
        
        with transaction.atomic():
            for assignment in assignments:
                try:
                    # Проверяем ячейку
                    storage_cell = Storage.objects.get(box_id=box_id, cell_id=assignment.cell_id)
                    
                    # Проверяем образец
                    sample = Sample.objects.get(id=assignment.sample_id)
                    
                    # Проверяем доступность ячейки
                    existing_sample = Sample.objects.filter(storage=storage_cell).exclude(comment_id=2).first()
                    if existing_sample:
                        errors.append(f'Ячейка {assignment.cell_id} уже занята образцом ID {existing_sample.id}')
                        continue
                    
                    # Проверяем, что образец не размещен в другой ячейке
                    if sample.storage and sample.comment_id != 2:
                        errors.append(f'Образец {assignment.sample_id} уже размещен в ячейке {sample.storage.cell_id}')
                        continue
                    
                    # Размещаем образец
                    sample.storage = storage_cell
                    sample.comment_id = None
                    sample.save()
                    
                    success_assignments.append({
                        'sample_id': sample.id,
                        'cell_id': assignment.cell_id,
                        'strain_code': sample.strain.short_code if sample.strain else None
                    })
                    
                    # Логируем размещение
                    log_change(
                        request=request,
                        content_type='Sample',
                        object_id=sample.id,
                        action='BULK_ASSIGN_CELL',
                        new_values={
                            'box_id': box_id,
                            'cell_id': assignment.cell_id,
                            'storage_id': storage_cell.id
                        },
                        comment='Bulk assignment of sample to cell'
                    )
                    
                except Storage.DoesNotExist:
                    errors.append(f'Ячейка {assignment.cell_id} в боксе {box_id} не найдена')
                except Sample.DoesNotExist:
                    errors.append(f'Образец с ID {assignment.sample_id} не найден')
                except Exception as e:
                    errors.append(f'Ошибка при размещении образца {assignment.sample_id} в ячейке {assignment.cell_id}: {str(e)}')
        
        return Response({
            'message': f'Массовое размещение завершено',
            'statistics': {
                'total_requested': len(assignments),
                'successful': len(success_assignments),
                'failed': len(errors)
            },
            'successful_assignments': success_assignments,
            'errors': errors
        })
    
    except Exception as e:
        logger.error(f"Error in bulk_assign_cells: {e}")
        return Response({
            'error': f'Ошибка при массовом размещении: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


MAX_IMAGE_SIZE_MB = 1
ALLOWED_IMAGE_TYPES = ["image/jpeg", "image/png"]


def _validate_uploaded_image(uploaded_file):
    if uploaded_file.content_type not in ALLOWED_IMAGE_TYPES:
        raise DjangoValidationError("Разрешены только JPEG и PNG файлы")

    if uploaded_file.size > MAX_IMAGE_SIZE_MB * 1024 * 1024:
        raise DjangoValidationError("Максимальный размер файла 1 МБ")


@api_view(["POST"])
def upload_sample_photos(request, sample_id):
    """Загружает одну или несколько фотографий для образца."""
    try:
        sample = Sample.objects.get(id=sample_id)
    except Sample.DoesNotExist:
        return Response({"error": "Образец не найден"}, status=status.HTTP_404_NOT_FOUND)

    if not request.FILES:
        return Response({"error": "Файлы не переданы"}, status=status.HTTP_400_BAD_REQUEST)

    created = []
    errors = []

    for file_obj in request.FILES.getlist("images"):
        try:
            _validate_uploaded_image(file_obj)
            photo = sample.photos.create(image=file_obj)
            created.append({"id": photo.id, "url": photo.image.url})
        except DjangoValidationError as e:
            errors.append(str(e))
        except Exception as e:
            logger.exception("Ошибка при загрузке фото: %s", e)
            errors.append(str(e))

    return Response({"created": created, "errors": errors})


@api_view(["DELETE"])
def delete_sample_photo(request, sample_id, photo_id):
    """Удаляет фотографию образца."""
    try:
        photo = SamplePhoto.objects.get(id=photo_id, sample_id=sample_id)
        photo.delete()
        return Response({"message": "Фото удалено"})
    except SamplePhoto.DoesNotExist:
        return Response({"error": "Фото не найдено"}, status=status.HTTP_404_NOT_FOUND)


